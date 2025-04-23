import customtkinter as ctk
from tkinter import messagebox, ttk, Toplevel, Text, Scrollbar
from datetime import datetime
import os
import logging
from openpyxl import Workbook, load_workbook
import json
import win32print
import win32api
from win32printing import Printer

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Constants
CONFIG_FILE = "app_config.json"
AUTOSAVE_INTERVAL = 300000  # 5 minutes in milliseconds

# Item list for dropdown
ITEM_LIST = [
    "MAIZE", "SOYABEAN", "LOBHA", "HULLI", "KADLI", "BLACK MOONG", 
    "CHAMAKI MOONG", "RAGI", "WHEAT", "RICE", "BILAJOLA", "BIJAPUR", 
    "CHS-5", "FEEDS", "KUSUBI", "SASAVI", "SAVI", "CASTER SEEDS", 
    "TOOR RED", "TOOR WHITE", "HUNASIBIKA", "SF", "AWARI"
]

def validate_float(value):
    """Validate if a string can be converted to float."""
    try:
        return float(value) if value.strip() else 0
    except ValueError:
        return 0

# Define font configurations
HEADER_FONT = ("Segoe UI", 24, "bold")
SUBHEADER_FONT = ("Segoe UI", 14)
LABEL_FONT = ("Segoe UI", 12)
ENTRY_FONT = ("Segoe UI", 12)
TABLE_HEADER_FONT = ("Segoe UI", 12, "bold")
TABLE_FONT = ("Segoe UI", 12)
BUTTON_FONT = ("Segoe UI", 12)

class InvoiceApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.load_config()
        self.setup_ui()
        # Removed the call to self.schedule_autosave() since it's not defined
        # Uncomment the next line if you plan to use autosave later
        # self.schedule_autosave()

    def load_config(self):
        """Load application configuration from file"""
        self.config = {
            "theme": "light",
            "window_size": "1200x800",
            "autosave": True
        }
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    self.config.update(json.load(f))
        except Exception as e:
            logging.error(f"Error loading config: {e}")

    def save_config(self):
        """Save application configuration to file"""
        try:
            with open(CONFIG_FILE, 'w') as f:
                json.dump(self.config, f)
        except Exception as e:
            logging.error(f"Error saving config: {e}")

    def setup_ui(self):
        """Initialize the user interface"""
        ctk.set_appearance_mode(self.config["theme"])
        ctk.set_default_color_theme("blue")
        
        self.title("GV Mahant Brothers - Invoice")
        self.geometry(self.config["window_size"])
        self.minsize(1000, 700)
        self.configure(padx=20, pady=20)

        self.current_mode = ctk.StringVar(value="Patti")
        self.rows = []
        self.row_counter = 0
        self.autosave_var = ctk.BooleanVar(value=self.config["autosave"])

        # Create tooltip label
        self.tooltip = ttk.Label(self, background="#ffffe0", relief="solid", borderwidth=1)
        self.tooltip_timer = None

        self.kata_amount_entry = None

        self.build_ui()

    def build_ui(self):
        # Spacer and Header
        ctk.CTkLabel(self, text="", height=1).pack()  # Spacer
        ctk.CTkLabel(self, text="|ಶ್ರೀ|", font=HEADER_FONT, text_color="#1976d2").pack()
        ctk.CTkLabel(self, text="G.V. Mahant Brothers", font=HEADER_FONT, text_color="#1976d2").pack()
        ctk.CTkLabel(self, text=datetime.now().strftime("%A, %d %B %Y %I:%M %p"), font=SUBHEADER_FONT).pack()

        # Mode navigation
        nav_frame = ctk.CTkFrame(self, fg_color="transparent")
        for i, mode in enumerate(["Patti", "Kata", "Barthe"]):
            rb = ctk.CTkRadioButton(nav_frame, text=mode, variable=self.current_mode, value=mode, command=self.switch_mode, font=LABEL_FONT)
            rb.grid(row=0, column=i, padx=20, pady=10)  # Increased spacing
        nav_frame.pack(pady=(30, 20))

        # Customer name
        customer_frame = ctk.CTkFrame(self, fg_color="transparent")
        ctk.CTkLabel(customer_frame, text="Customer Name:", font=LABEL_FONT).pack(side="left", padx=(0, 10))
        self.customer_entry = ctk.CTkEntry(customer_frame, width=400, font=ENTRY_FONT, height=35)  # Increased width and height
        self.customer_entry.pack(side="left")
        customer_frame.pack(pady=(20, 30))

        # Table container Frame (Define it here)
        self.table_frame = ctk.CTkFrame(self, fg_color="transparent")
        # self.rows is already initialized in __init__
        # self.create_table_headers() # Create headers *after* frames are packed

        # --- Buttons and Total Frame --- (Define and PACK this FIRST)
        self.bottom_frame = ctk.CTkFrame(self, fg_color="transparent")
        # Pack bottom frame against the window bottom, with some padding
        self.bottom_frame.pack(fill="x", pady=(10, 10), side="bottom", anchor="sw") # pady can be adjusted

        # --- Left side: Action Buttons (inside bottom_frame) ---
        left_buttons_frame = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        left_buttons_frame.pack(side="left", padx=(0, 20)) 
        # (Button creation remains the same)
        button_params = {"font": BUTTON_FONT, "width": 120, "height": 35}
        ctk.CTkButton(left_buttons_frame, text="Add Row", command=self.add_row, **button_params).pack(side="left", padx=5)
        ctk.CTkButton(left_buttons_frame, text="Clear", command=self.clear_rows, **button_params).pack(side="left", padx=5)
        ctk.CTkButton(left_buttons_frame, text="Save", command=self.save_to_excel, **button_params).pack(side="left", padx=5) 
        ctk.CTkButton(left_buttons_frame, text="Print", command=self.show_print_preview, **button_params).pack(side="left", padx=5) 

        # --- Right side: Total and Optional Kata Field (inside bottom_frame) ---
        right_total_frame = ctk.CTkFrame(self.bottom_frame, fg_color="transparent")
        right_total_frame.pack(side="right")
        self.kata_field_frame = ctk.CTkFrame(right_total_frame, fg_color="transparent")
        self.kata_field_frame.pack(side="left", padx=(0, 10)) 
        self.total_label = ctk.CTkLabel(right_total_frame, text="Amount: ₹0.00", font=("Segoe UI", 16, "bold"))
        self.total_label.pack(side="left") 
        # --- End Buttons and Total Frame Setup ---

        # --- Now PACK the Table Frame ABOVE the bottom_frame ---
        # Make it fill available space horizontally and vertically, add padding
        self.table_frame.pack(fill="both", expand=True, padx=0, pady=(0, 10)) # pady adds space above bottom_frame

        # Create initial table content AFTER packing frames
        self.create_table_headers() 
        self.add_row() 
        self.switch_mode() 

    def create_table_headers(self):
        # Remove all widgets in table_frame
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        # self.rows should be cleared in switch_mode or clear_rows, not here
        # self.rows.clear() 

        headers = []
        if self.current_mode.get() == "Patti":
            headers = ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount"]
        elif self.current_mode.get() == "Kata":
            headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount"]
        elif self.current_mode.get() == "Barthe":
            headers = ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount"]

        # Store headers for reference if needed elsewhere (optional)
        self._current_headers = headers 

        for i, h in enumerate(headers):
            ctk.CTkLabel(
                self.table_frame,
                text=h,
                font=TABLE_HEADER_FONT,
                text_color="white",
                fg_color="#1976d2", # Or use theme color
                corner_radius=5,
                # width=100, # Let grid weight handle width
                height=35,  # Added height
                anchor="center"
            ).grid(row=0, column=i, sticky="nsew", padx=2, pady=2)
            self.table_frame.grid_columnconfigure(i, weight=1) # Make columns resize

    def switch_mode(self):
        # --- Clear Kata field if it exists ---
        for widget in self.kata_field_frame.winfo_children():
            widget.destroy()
        self.kata_amount_entry = None # Reset the variable
        # --- End Clear Kata field ---

        # Recreate headers which also clears the table frame's rows
        self.create_table_headers()
        # Clear the logical rows list
        self.rows.clear() 
        self.add_row() # Add a new blank row for the new mode

        # --- Add Kata field if mode is Kata ---
        if self.current_mode.get() == "Kata":
            kata_label = ctk.CTkLabel(self.kata_field_frame, text="Kata:", font=LABEL_FONT)
            kata_label.pack(side="left", padx=(0, 5))
            
            self.kata_amount_entry = ctk.CTkEntry(
                self.kata_field_frame, 
                font=ENTRY_FONT,
                height=35,
                width=120
            )
            self.kata_amount_entry.pack(side="left")
            # Add default value '0'
            self.kata_amount_entry.insert(0, "0") 
            # Bind update on key release
            self.kata_amount_entry.bind("<KeyRelease>", self.update_amounts)
        # --- End Add Kata field ---

        self.update_amounts() # Recalculate total

    def add_row(self):
        mode = self.current_mode.get()
        # Determine the number of entry fields based on mode (excluding the Amount label)
        if mode == "Patti":
            num_entry_fields = 5 # Item, Pkt, Qty, Rate, Hamali
        elif mode == "Kata":
            num_entry_fields = 5 # Item, Net, Less, Rate, Hamali Rate
        elif mode == "Barthe":
            num_entry_fields = 6 # Item, Pkt, Wt, +/-, Rate, Hamali
        else: # Default or fallback
            num_entry_fields = 5 

        entries = []
        # Start grid row below the header (row 0) and subsequent rows
        row_idx = len(self.rows) + 1 
        
        # --- Create Item Dropdown (ttk.Combobox for the first column) ---
        item_dropdown = ttk.Combobox(
            self.table_frame,
            values=ITEM_LIST,
            font=TABLE_FONT,
            # height=10, # Height isn't directly settable like CTkEntry
            state="readonly" # Prevent typing custom values
        )
        item_dropdown.grid(row=row_idx, column=0, padx=2, pady=2, sticky="nsew")
        # Bind selection event to update amounts
        item_dropdown.bind("<<ComboboxSelected>>", self.update_amounts) 
        self.table_frame.grid_columnconfigure(0, weight=1) # Ensure column resizes
        entries.append(item_dropdown)
        # --- End Item Dropdown ---

        # --- Create remaining entry widgets (starting from column 1) ---
        for i in range(1, num_entry_fields): # Start loop from 1
            entry = ctk.CTkEntry(
                self.table_frame,
                font=TABLE_FONT,
                justify="center",
                height=35  # Increased height
            )
            entry.grid(row=row_idx, column=i, padx=2, pady=2, sticky="nsew")
            # Update amounts whenever a key is released in any entry
            entry.bind("<KeyRelease>", self.update_amounts) # Simplified binding
            self.table_frame.grid_columnconfigure(i, weight=1) # Ensure column resizes
            entries.append(entry)
        # --- End remaining entries ---

        # Create the Amount label for this row (last column)
        amount_label = ctk.CTkLabel(
            self.table_frame,
            text="₹0.00",
            font=TABLE_FONT,
            anchor="e", # Align text to the right
            height=35  # Increased height
        )
        # Place it in the column after the last entry field
        amount_label.grid(row=row_idx, column=num_entry_fields, padx=2, pady=2, sticky="nsew")
        self.table_frame.grid_columnconfigure(num_entry_fields, weight=1) # Ensure column resizes
        
        # Append the amount label to the list of widgets for this row
        entries.append(amount_label) 

        # Store the row index and the list of widgets (dropdown + entries + amount label)
        self.rows.append({"row_index": row_idx, "widgets": entries})
        # Don't call update_amounts here, it's called by events or explicitly elsewhere
        # self.update_amounts()

    def clear_rows(self):
        # Destroy all widget rows (skip header row 0)
        for widget in self.table_frame.winfo_children():
            grid_info = widget.grid_info()
            # Check if 'row' key exists and if row > 0
            if grid_info and grid_info.get('row', 0) > 0:
                widget.destroy()
        
        # Clear the logical list of rows
        self.rows.clear()
        # Reset customer entry (optional, consider if needed)
        # self.customer_entry.delete(0, 'end') 
        self.add_row() # Add a single blank row back
        self.update_amounts() # Update total (should be 0)

    def update_amounts(self, event=None): # Accept event argument from binding
        logging.debug("Updating amounts for all rows")
        total = 0.0
        mode = self.current_mode.get()

        # Calculate sum of row amounts
        for row_data in self.rows:
            widgets = row_data["widgets"]
            amount = 0.0 # Default amount
            try:
                if mode == "Patti":
                    # Item [0], Pkt [1], Qty [2], Rate [3], Hamali [4], AmountLabel [5]
                    if len(widgets) > 4:
                        qty = validate_float(widgets[2].get())
                        rate = validate_float(widgets[3].get())
                        pkt = validate_float(widgets[1].get())
                        hamali = validate_float(widgets[4].get())
                        amount = (qty * rate) + (pkt * hamali)
                elif mode == "Kata":
                    # Item [0], Net [1], Less% [2], Rate [3], HamaliRate [4], AmountLabel [5]
                     if len(widgets) > 4:
                        net = validate_float(widgets[1].get())
                        less = validate_float(widgets[2].get())
                        final_wt = net * (1 - less / 100.0) if less < 100 else 0.0
                        rate = validate_float(widgets[3].get())
                        hamali_rate = validate_float(widgets[4].get())
                        # Calculate packets based on net weight (e.g., if 60kg/packet)
                        # This logic might need adjustment based on actual use case
                        packets = int(net / 60) if net > 0 else 0 
                        amount = (final_wt * rate) + (packets * hamali_rate)
                elif mode == "Barthe":
                    # Item [0], Pkt [1], Wt/Pkt [2], +/- Adj [3], Rate [4], Hamali/Pkt [5], AmountLabel [6]
                     if len(widgets) > 5:
                        pkt = validate_float(widgets[1].get())
                        wt_per_pkt = validate_float(widgets[2].get())
                        adj = validate_float(widgets[3].get()) # Adjustment weight
                        rate = validate_float(widgets[4].get()) # Rate per kg
                        hamali_per_pkt = validate_float(widgets[5].get())
                        total_qty = (pkt * wt_per_pkt) + adj
                        amount = (total_qty * rate) + (pkt * hamali_per_pkt)
                
                # Update the amount label for the current row
                if len(widgets) > 0:
                    widgets[-1].configure(text=f"₹{amount:.2f}")
                total += amount

            except IndexError:
                 logging.error(f"Index error calculating amount for row. Widgets: {len(widgets)}")
            except Exception as e:
                logging.error(f"Error calculating amount: {e}")
                if len(widgets) > 0:
                     widgets[-1].configure(text="₹Error") # Indicate error on the row

        # --- Add Kata Amount if applicable ---
        kata_amount = 0.0
        if mode == "Kata" and self.kata_amount_entry:
            try:
                kata_amount = validate_float(self.kata_amount_entry.get())
                # Add visual feedback for invalid input (optional)
                if self.kata_amount_entry.get().strip() and kata_amount == 0 and self.kata_amount_entry.get() != '0':
                     self.kata_amount_entry.configure(fg_color="pink")
                else:
                     # Reset color on valid input
                     self.kata_amount_entry.configure(fg_color=ctk.ThemeManager.theme["CTkEntry"]["fg_color"]) 
            except Exception as e:
                logging.error(f"Error reading Kata amount: {e}")
                # Maybe provide visual feedback on error
                self.kata_amount_entry.configure(fg_color="pink")
        
        total += kata_amount # Add validated Kata amount to total
        # --- End Add Kata Amount ---

        self.total_label.configure(text=f"Amount: ₹{total:.2f}")

    def save_to_excel(self): # Renamed from save_invoice for clarity
        try:
            # --- Construct the full path to the Documents folder ---
            try:
                # Get user's home directory
                home_dir = os.path.expanduser("~") 
                # Create the full path to the Documents folder
                documents_path = os.path.join(home_dir, "Documents")
                
                # Ensure the Documents directory exists, create if not
                os.makedirs(documents_path, exist_ok=True) 
                
                # Create the filename based on the current date
                date_str = datetime.now().strftime('%Y-%m-%d')
                base_filename = f"Invoice_{date_str}.xlsx"
                
                # Combine documents path and filename
                full_save_path = os.path.join(documents_path, base_filename)
                
                logging.info(f"Target save path: {full_save_path}")

            except Exception as path_e:
                 logging.error(f"Error determining save path: {path_e}")
                 messagebox.showerror("Path Error", f"Could not determine the Documents folder path.\nError: {path_e}")
                 return # Stop if we can't get the path

            # --- Get Invoice Data ---
            customer = self.customer_entry.get().strip() or "Unknown Customer"
            mode = self.current_mode.get()
            
            headers = getattr(self, '_current_headers', []) 
            if not headers: # Fallback
                 # (Fallback header logic remains the same)
                 if mode == "Patti":
                     headers = ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount"]
                 elif mode == "Kata":
                     headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount"]
                 elif mode == "Barthe":
                     headers = ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount"]
                 else: 
                     headers = ["Col1", "Col2", "Col3", "Col4", "Col5", "Col6", "Amount"]

            data_rows = []
            for row_data in self.rows:
                 widgets = row_data["widgets"]
                 # Use .get() for Combobox and CTkEntry, .cget() for CTkLabel
                 row_values = []
                 for w in widgets:
                      if isinstance(w, (ctk.CTkEntry, ttk.Combobox)):
                           row_values.append(w.get())
                      elif isinstance(w, ctk.CTkLabel):
                           row_values.append(w.cget("text"))
                      else: # Fallback for unexpected widget types
                           row_values.append("") 

                 if row_values:
                     # Clean amount (assuming it's the last value)
                     row_values[-1] = row_values[-1].replace('₹', '').replace('Error', '0')
                     # Only include rows with actual item data (check first column)
                     if row_values[0] and row_values[0].strip(): 
                          data_rows.append(row_values)

            if not data_rows:
                messagebox.showwarning("No Data", "No data entered to save.")
                return

            # --- Excel Writing Logic ---
            if os.path.exists(full_save_path): # Check existence using the full path
                try:
                    wb = load_workbook(full_save_path)
                    if mode in wb.sheetnames:
                         ws = wb[mode]
                    else:
                         ws = wb.create_sheet(title=mode)
                         ws.append(["Timestamp", "Customer"] + headers) 
                except Exception as e:
                     logging.error(f"Error loading workbook '{full_save_path}', creating new one: {e}")
                     wb = Workbook()
            ws = wb.active
            ws.title = mode 
            ws.append(["Timestamp", "Customer"] + headers)

            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            for row in data_rows:
                ws.append([timestamp, customer] + row)

            # --- Save Workbook ---
            try:
                # Save using the full path to the Documents folder
                wb.save(full_save_path) 
                logging.info(f"Successfully saved invoice data to {full_save_path} (Sheet: {mode})")
                # Update message box to show the full path
                messagebox.showinfo("Saved", f"Invoice data saved to:\n{full_save_path}\n(Sheet: {mode})") 
            except PermissionError:
                error_msg = f"Cannot save '{base_filename}'.\nThe file might be open in Excel.\n\nLocation: {documents_path}"
                logging.error(error_msg)
                messagebox.showerror("Permission Error", error_msg)
            except Exception as e:
                error_msg = f"Error saving Excel file to:\n{full_save_path}\n\nError: {str(e)}"
                logging.error(error_msg)
                messagebox.showerror("Save Error", error_msg)

        except Exception as e:
            # Catch any other unexpected errors during the process
            error_msg = f"Unexpected error during save operation: {str(e)}"
            logging.exception(error_msg) # Log the full traceback
            messagebox.showerror("Error", error_msg)

    def generate_print_content(self):
        """Generates the formatted string list for printing/preview."""
        lines = []
        customer = self.customer_entry.get().strip() or "N/A"
        mode = self.current_mode.get()
        max_width = 40 # Keep max width for typical thermal printers

        # --- Header ---
        lines.extend([
            "|ಶ್ರೀ|".center(max_width),
            "",
            "G.V. Mahant Brothers".center(max_width),
            datetime.now().strftime("%d-%b-%Y %H:%M").center(max_width),
            "-" * max_width,
            f"Customer: {customer}",
            f"Mode: {mode}",
            "-" * max_width,
        ])

        # --- Column Headers ---
        # Revised widths and spacing for better alignment within 40 chars
        header_fmt = ""
        if mode == "Patti":
            # Item(8) Pkt(5) Qty(6) Rate(6) Ham(5) Amt(7) = 37 chars + 5 spaces = 42 -> Adjust
            # Try: Item(8) Pkt(5) Qty(5) Rate(6) Ham(5) Amt(7) = 36 chars + 5 spaces = 41 -> Adjust
            # Try: Item(8) Pkt(4) Qty(5) Rate(6) Ham(5) Amt(7) = 35 chars + 5 spaces = 40
            header_fmt = "{:<8} {:>4} {:>5} {:>6} {:>5} {:>7}" 
            headers = ["Item", "Pkt", "Qty", "Rate", "Hamali", "Amount"]
        elif mode == "Kata":
            # Item(8) Net(7) Less(5) Rate(6) Ham(5) Amt(7) = 38 chars + 5 spaces = 43 -> Adjust
            # Try: Item(8) Net(6) Less(4) Rate(6) Ham(5) Amt(7) = 36 chars + 5 spaces = 41 -> Adjust
            # Try: Item(8) Net(6) Less(4) Rate(5) Ham(5) Amt(7) = 35 chars + 5 spaces = 40
            header_fmt = "{:<8} {:>6} {:>4} {:>5} {:>5} {:>7}"
            headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali", "Amount"]
        elif mode == "Barthe":
             # Item(7) Pkt(5) Wt(5) +/- (5) Rate(6) Ham(5) Amt(7) = 40 chars + 6 spaces = 46 -> Adjust
             # Try: Item(7) Pkt(4) Wt(5) +/- (4) Rate(5) Ham(5) Amt(7) = 37 chars + 6 spaces = 43 -> Adjust
             # Try: Item(7) Pkt(4) Wt(4) +/- (4) Rate(5) Ham(5) Amt(7) = 36 chars + 6 spaces = 42 -> Adjust
             # Try: Item(7) Pkt(4) Wt(4) +/- (4) Rate(5) Ham(4) Amt(7) = 35 chars + 6 spaces = 41 -> Adjust
             # Try: Item(6) Pkt(4) Wt(4) +/- (4) Rate(5) Ham(4) Amt(7) = 34 chars + 6 spaces = 40
             header_fmt = "{:<6} {:>4} {:>4} {:>4} {:>5} {:>4} {:>7}"
             headers = ["Item", "Pkt", "Wt", "+/-", "Rate", "Hamali", "Amount"]
        
        if header_fmt: # Check if format string was set
             lines.append(header_fmt.format(*headers)) # Use tuple unpacking
        else:
             lines.append("Error: Mode not recognized for printing.") # Fallback
        
        lines.append("-" * max_width)

        # --- Data Rows ---
        for row_data in self.rows:
            widgets = row_data["widgets"]
            # Extract values using .get() or .cget()
            row_values = []
            for w in widgets:
                if isinstance(w, (ctk.CTkEntry, ttk.Combobox)):
                    row_values.append(w.get().strip()) # Get and strip whitespace
                elif isinstance(w, ctk.CTkLabel):
                     # Get text, remove currency symbol and errors for formatting
                    text = w.cget("text").replace('₹', '').replace('Error', '0').strip()
                    row_values.append(text)
                else: 
                    row_values.append("") 

            # Skip if item name (first field) is empty or only whitespace
            if not row_values or not row_values[0]:
                 continue
            
            # Use the same format string as headers for data alignment
            # No truncation (.N) applied here to avoid cutting numbers, 
            # but extremely long entries might still misalign.
            try:
                if mode == "Patti" and len(row_values) >= 6:
                     lines.append(header_fmt.format(
                         row_values[0][:8],  # Item (limit length)
                         row_values[1],      # Pkt
                         row_values[2],      # Qty
                         row_values[3],      # Rate
                         row_values[4],      # Hamali
                         row_values[5]       # Amount (already cleaned)
                     ))
                elif mode == "Kata" and len(row_values) >= 6:
                      lines.append(header_fmt.format(
                         row_values[0][:8],  # Item (limit length)
                         row_values[1],      # Net Wt
                         row_values[2],      # Less%
                         row_values[3],      # Rate
                         row_values[4],      # Hamali
                         row_values[5]       # Amount (already cleaned)
                      ))
                elif mode == "Barthe" and len(row_values) >= 7:
                     lines.append(header_fmt.format(
                         row_values[0][:6],  # Item (limit length)
                         row_values[1],      # Pkt
                         row_values[2],      # Wt
                         row_values[3],      # +/-
                         row_values[4],      # Rate
                         row_values[5],      # Hamali
                         row_values[6]       # Amount (already cleaned)
                     ))
            except IndexError:
                 lines.append("Error formatting row data...")
                 logging.warning(f"IndexError formatting print data row: {row_values}")
            except Exception as fmt_e:
                 lines.append(f"Fmt Error: {fmt_e}")
                 logging.warning(f"Exception formatting print data row: {fmt_e} - {row_values}")

        # --- Add Kata Amount if applicable ---
        if mode == "Kata" and self.kata_amount_entry:
             kata_val_str = self.kata_amount_entry.get().strip()
             kata_amount = validate_float(kata_val_str)
             # Format nicely, right-aligned
             lines.append(f"Kata Amount: {kata_amount:>10.2f}".rjust(max_width))
        # --- End Add Kata Amount ---

        # --- Footer ---
        lines.extend([
            "-" * max_width,
            # Right-align the total amount string within the max_width
            f"Total Amount: {self.total_label.cget('text')}".rjust(max_width), 
            "-" * max_width,
            "",
            # Optional Kannada text - ensure your printer supports it
            # "ನಾನು ಎಲ್ಲವೂ ಸರಿಯಾಗಿದೆ ಎಂದು ಪರಿಶೀಲಿಸಿದ್ದೇನೆ.".center(max_width), 
            "",
            "_" * max_width,
            "Customer Signature".center(max_width),
            "\n\n"  # Extra lines for paper feed/cut
        ])

        # Add printer cut command (ESC/POS standard)
        lines.append(chr(27) + chr(105)) # Full cut
        # lines.append(chr(27) + chr(109)) # Partial cut

        return lines

    def save_for_print(self):
        """Prints the generated content to the default printer."""
        try:
            printer_name = win32print.GetDefaultPrinter()
            logging.info(f"Attempting to print to default printer: {printer_name}")
            
            lines = self.generate_print_content()
            print_content = "\n".join(lines)
            
            # Ensure content is encoded correctly, try 'cp437' for older ESC/POS
            # or utf-8 if your printer supports it. 'cp949' or others might be needed for specific languages.
            print_bytes = print_content.encode('cp437', errors='replace') 

            # Use win32print for direct RAW printing
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                # Job name "Invoice", Datatype "RAW"
                hJob = win32print.StartDocPrinter(hPrinter, 1, ("Invoice", None, "RAW"))
                try:
                    win32print.StartPagePrinter(hPrinter)
                    win32print.WritePrinter(hPrinter, print_bytes)
                    win32print.EndPagePrinter(hPrinter)
                finally:
                    win32print.EndDocPrinter(hPrinter)
            finally:
                win32print.ClosePrinter(hPrinter)

            logging.info("Invoice successfully sent to printer.")
            messagebox.showinfo("Success", "Invoice sent to printer!")

        except Exception as e:
            error_msg = f"Error printing invoice: {str(e)}"
            logging.error(error_msg)
            # Detailed error for debugging might be helpful
            # import traceback
            # logging.error(traceback.format_exc()) 
            messagebox.showerror("Print Error", f"Could not print to {printer_name}.\nCheck printer connection and status.\n\nError: {e}")


    def show_print_preview(self):
        """Shows a Toplevel window with a preview of the print output."""
        try:
            preview = ctk.CTkToplevel(self)
            preview.title("Print Preview")
            preview.geometry("450x600") # Slightly wider for better view
            preview.transient(self) # Keep preview on top of main window
            preview.grab_set()  # Make the window modal

            # Center the preview window relative to the main app
            app_x = self.winfo_x()
            app_y = self.winfo_y()
            app_w = self.winfo_width()
            app_h = self.winfo_height()
            pre_w = 450
            pre_h = 600
            x = app_x + (app_w - pre_w) // 2
            y = app_y + (app_h - pre_h) // 2
            preview.geometry(f"{pre_w}x{pre_h}+{x}+{y}")

            # --- Preview Content Area ---
            # Use a CTkTextbox within a ScrollableFrame for better handling
            scroll_frame = ctk.CTkScrollableFrame(preview, label_text="Preview")
            scroll_frame.pack(fill="both", expand=True, padx=10, pady=(10, 0))

            preview_text = ctk.CTkTextbox(
                scroll_frame, 
                font=("Courier New", 10), # Monospace font for alignment
                wrap="none" # Prevent wrapping to see true line breaks
            )
            preview_text.pack(fill="both", expand=True)
            
            # Generate print content and display it
            lines = self.generate_print_content()
            # Join lines, but remove the final cut command for preview
            preview_content = "\n".join(lines[:-1]) if lines else "" 
            
            preview_text.insert("1.0", preview_content)
            preview_text.configure(state="disabled")  # Make read-only

            # --- Buttons Frame ---
            button_frame = ctk.CTkFrame(preview, fg_color="transparent")
            button_frame.pack(fill="x", padx=10, pady=10)
            
            # Center buttons using grid
            button_frame.grid_columnconfigure(0, weight=1)
            button_frame.grid_columnconfigure(1, weight=1)

            ctk.CTkButton(
                button_frame,
                text="Print",
                # Lambda calls destroy first, then the print function
                command=lambda: [preview.destroy(), self.save_for_print()], 
                width=120
            ).grid(row=0, column=0, padx=5, pady=5, sticky="ew")

            ctk.CTkButton(
                button_frame,
                text="Close",
                command=preview.destroy,
                width=120
            ).grid(row=0, column=1, padx=5, pady=5, sticky="ew")
            
            preview.after(100, preview.lift) # Ensure it's raised above main window

        except Exception as e:
            error_msg = f"Error generating print preview: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("Preview Error", error_msg)
            if 'preview' in locals() and preview.winfo_exists():
                 preview.destroy() # Close broken preview window

    # Uncomment this if you want to implement autosave later
    # def schedule_autosave(self):
    #     """Schedule an automatic save every 5 minutes"""
    #     self.after(AUTOSAVE_INTERVAL, self.autosave)  # AUTOSAVE_INTERVAL is 300000 ms (5 minutes)
    #
    # def autosave(self):
    #     """Perform autosave task (e.g., save invoice)"""
    #     self.save_invoice()
    #     logging.info("Autosave completed.")
    #     # Reschedule the next autosave
    #     self.schedule_autosave()

if __name__ == "__main__":
    app = InvoiceApp()
    app.mainloop()
