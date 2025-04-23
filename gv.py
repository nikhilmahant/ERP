import customtkinter as ctk
from tkinter import messagebox
from datetime import datetime
import os
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import webview
import threading
import tempfile
import sys
import webbrowser
from tkinter import ttk
import tkinterweb
import win32print
import win32api
from win32printing import Printer

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def validate_float(value):
    """Validate if a string can be converted to float."""
    try:
        return float(value) if value.strip() else 0
    except ValueError:
        return 0

# Define font configurations
HEADER_FONT = ("Segoe UI", 24, "bold")
SUBHEADER_FONT = ("Segoe UI", 14)
LABEL_FONT = ("Segoe UI", 12)
ENTRY_FONT = ("Segoe UI", 12)
TABLE_HEADER_FONT = ("Segoe UI", 12, "bold")
TABLE_FONT = ("Segoe UI", 12)
BUTTON_FONT = ("Segoe UI", 12)

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class InvoiceApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.title("G.V. Mahant Brothers - Invoice")
        self.geometry("1200x800")  # Increased window size
        self.configure(padx=20, pady=20)  # Increased padding

        self.current_mode = ctk.StringVar(value="Patti")
        self.rows = []
        self.row_counter = 0

        self.build_ui()

    def build_ui(self):
        # Spacer and Header
        ctk.CTkLabel(self, text="", height=1).pack()  # Spacer
        ctk.CTkLabel(self, text="|ಶ್ರೀ|", font=HEADER_FONT, text_color="#1976d2").pack()
        ctk.CTkLabel(self, text="G.V. Mahant Brothers", font=HEADER_FONT, text_color="#1976d2").pack()
        ctk.CTkLabel(self, text=datetime.now().strftime("%A, %d %B %Y %I:%M %p"), font=SUBHEADER_FONT).pack()

        # Mode navigation
        nav_frame = ctk.CTkFrame(self, fg_color="transparent")
        for i, mode in enumerate(["Patti", "Kata", "Barthe"]):
            rb = ctk.CTkRadioButton(nav_frame, text=mode, variable=self.current_mode, value=mode, command=self.switch_mode, font=LABEL_FONT)
            rb.grid(row=0, column=i, padx=20, pady=10)  # Increased spacing
        nav_frame.pack(pady=(30, 20))

        # Customer name
        customer_frame = ctk.CTkFrame(self, fg_color="transparent")
        ctk.CTkLabel(customer_frame, text="Customer Name:", font=LABEL_FONT).pack(side="left", padx=(0, 10))
        self.customer_entry = ctk.CTkEntry(customer_frame, width=400, font=ENTRY_FONT, height=35)  # Increased width and height
        self.customer_entry.pack(side="left")
        customer_frame.pack(pady=(20, 30))

        # Table container
        self.table_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.table_frame.pack(fill="x", pady=(0, 20))
        self.rows = []
        self.create_table_headers()

        # Buttons and total
        bottom_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_params = {"font": BUTTON_FONT, "width": 120, "height": 35}  # Increased button size
        ctk.CTkButton(bottom_frame, text="Add Row", command=self.add_row, **button_params).pack(side="left", padx=10)
        ctk.CTkButton(bottom_frame, text="Clear", command=self.clear_rows, **button_params).pack(side="left", padx=10)
        ctk.CTkButton(bottom_frame, text="Save", command=self.save_to_excel, **button_params).pack(side="left", padx=10)
        ctk.CTkButton(bottom_frame, text="Print", command=self.show_print_preview, **button_params).pack(side="left", padx=10)
        self.total_label = ctk.CTkLabel(bottom_frame, text="Amount: ₹0.00", font=("Segoe UI", 16, "bold"))
        self.total_label.pack(side="right")
        bottom_frame.pack(fill="x", pady=(20, 0))

        self.add_row()

    def create_table_headers(self):
        # Remove all widgets in table_frame
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        self.rows.clear()

        headers = []
        if self.current_mode.get() == "Patti":
            headers = ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount"]
        elif self.current_mode.get() == "Kata":
            headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount"]
        elif self.current_mode.get() == "Barthe":
            headers = ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount"]

        for i, h in enumerate(headers):
            ctk.CTkLabel(
                self.table_frame,
                text=h,
                font=TABLE_HEADER_FONT,
                text_color="white",
                fg_color="#1976d2",
                corner_radius=5,
                width=100,  # Increased width
                height=35,  # Added height
                anchor="center"
            ).grid(row=0, column=i, sticky="nsew", padx=2, pady=2)
            self.table_frame.grid_columnconfigure(i, weight=1)

    def switch_mode(self):
        # Remove all widgets (headers + data rows)
        for widget in self.table_frame.winfo_children():
            widget.destroy()
        self.rows.clear()
        self.create_table_headers()
        self.add_row()
        self.update_amounts()

    def add_row(self):
        mode = self.current_mode.get()
        if mode == "Patti":
            num_fields = 5
        elif mode == "Kata":
            num_fields = 5
        elif mode == "Barthe":
            num_fields = 6
        else:
            num_fields = 5

        entries = []
        row_idx = len(self.rows) + 1
        bg_even = "#e3f2fd"
        bg_odd = "#ffffff"
        row_bg = bg_even if row_idx % 2 == 0 else bg_odd
        
        for i in range(num_fields):
            entry = ctk.CTkEntry(
                self.table_frame,
                font=TABLE_FONT,
                justify="center",
                height=35  # Increased height
            )
            entry.grid(row=row_idx, column=i, padx=2, pady=2, sticky="nsew")
            entry.bind("<KeyRelease>", lambda e, ent=entry: self.update_amounts())
            self.table_frame.grid_columnconfigure(i, weight=1)
            entries.append(entry)

        amount_label = ctk.CTkLabel(
            self.table_frame,
            text="₹0.00",
            font=TABLE_FONT,
            anchor="e",
            height=35  # Increased height
        )
        amount_label.grid(row=row_idx, column=num_fields, padx=2, pady=2, sticky="nsew")
        self.table_frame.grid_columnconfigure(num_fields, weight=1)
        entries.append(amount_label)

        self.rows.append((row_idx, entries))
        self.update_amounts()

    def clear_rows(self):
        # Remove all widgets except headers
        for widget in self.table_frame.winfo_children():
            info = widget.grid_info()
            if info['row'] != 0:
                widget.destroy()
        self.rows.clear()
        self.row_counter = 0
        self.add_row()
        self.update_amounts()

    def update_amounts(self):
        logging.debug("Updating amounts for all rows")
        total = 0
        for _, widgets in self.rows:
            try:
                if self.current_mode.get() == "Patti":
                    qty = validate_float(widgets[2].get())
                    rate = validate_float(widgets[3].get())
                    pkt = validate_float(widgets[1].get())
                    hamali = validate_float(widgets[4].get())
                    amount = qty * rate + pkt * hamali
                elif self.current_mode.get() == "Kata":
                    net = validate_float(widgets[1].get())
                    less = validate_float(widgets[2].get())
                    final = net * (1 - less / 100)
                    rate = validate_float(widgets[3].get())
                    hamali = validate_float(widgets[4].get())
                    packets = int(net / 60) if net else 0
                    amount = final * rate + packets * hamali
                elif self.current_mode.get() == "Barthe":
                    pkt = validate_float(widgets[1].get())
                    wt = validate_float(widgets[2].get())
                    adj = validate_float(widgets[3].get())
                    rate = validate_float(widgets[4].get())
                    hamali = validate_float(widgets[5].get())
                    qty = pkt * wt + adj
                    amount = qty * rate + pkt * hamali
                else:
                    amount = 0
            except Exception as ex:
                logging.error(f"Calculation error: {ex}")
                amount = 0

            widgets[-1].configure(text=f"₹{amount:.2f}")
            total += amount

        self.total_label.configure(text=f"Amount: ₹{total:.2f}")

    def save_to_excel(self):
        try:
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f"Invoice_{date_str}.xlsx"
            customer = self.customer_entry.get().strip() or "Unknown Customer"
            mode = self.current_mode.get()
            
            # Determine headers based on mode
            headers = []
            if mode == "Patti":
                headers = ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount"]
            elif mode == "Kata":
                headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount"]
            elif mode == "Barthe":
                headers = ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount"]
            else:
                headers = ["Item", "Field1", "Field2", "Field3", "Field4", "Field5", "Amount"]

            # Collect data rows
            data_rows = []
            for _, widgets in self.rows:
                row = []
                for w in widgets[:-1]:
                    row.append(w.get())
                row.append(widgets[-1].cget("text").replace('₹',''))
                if any(cell.strip() for cell in row):
                    data_rows.append(row)

            if not data_rows:
                messagebox.showwarning("No Data", "No data to save.")
                return

            try:
                if os.path.exists(filename):
                    wb = load_workbook(filename)
                    ws = wb.active
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(["Date", "Customer", "Mode"] + headers)

                for row in data_rows:
                    ws.append([datetime.now().strftime('%Y-%m-%d %H:%M'), customer, mode] + row)

                wb.save(filename)
                logging.info(f"Successfully saved invoice data to {filename}")
                messagebox.showinfo("Saved", f"Invoice data saved to {filename}")

            except PermissionError:
                error_msg = f"Cannot access {filename}. The file might be open in another program."
                logging.error(error_msg)
                messagebox.showerror("Permission Error", error_msg)
            except Exception as e:
                error_msg = f"Error saving to Excel: {str(e)}"
                logging.error(error_msg)
                messagebox.showerror("Save Error", error_msg)

        except Exception as e:
            error_msg = f"Unexpected error: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("Error", error_msg)

    def save_for_print(self):
        """Print the invoice directly to the printer."""
        try:
            # Get the default printer
            printer_name = win32print.GetDefaultPrinter()
            
            # Prepare the text content
            lines = []
            
            # Header
            lines.extend([
                "          |ಶ್ರೀ|",
                "",
                "    G.V. Mahant Brothers",
                datetime.now().strftime("%A, %d %B %Y %I:%M %p"),
                "-" * 40,
                f"Customer: {self.customer_entry.get().strip()}",
                f"Mode: {self.current_mode.get()}",
                "-" * 40,
            ])
            
            # Add headers based on mode
            if self.current_mode.get() == "Patti":
                lines.append("Item     Pkt   Qty   Rate  Ham    Amt")
            elif self.current_mode.get() == "Kata":
                lines.append("Item     Net   Less  Rate  Ham    Amt")
            elif self.current_mode.get() == "Barthe":
                lines.append("Item     Pkt   Wt    +/-   Rate   Amt")
            
            lines.append("-" * 40)
            
            # Add rows
            for _, widgets in self.rows:
                if any(w.get().strip() for w in widgets[:-1]):
                    row_text = []
                    # Item name (limited to 8 chars)
                    row_text.append(f"{widgets[0].get()[:8]:<8}")
                    
                    # Other fields with appropriate spacing
                    for w in widgets[1:-1]:
                        row_text.append(f"{w.get()[:5]:>5}")
                    
                    # Amount
                    amount = widgets[-1].cget("text").replace("₹", "")
                    row_text.append(f"{amount:>6}")
                    
                    lines.append("".join(row_text))
            
            lines.extend([
                "-" * 40,
                f"Total Amount: {self.total_label.cget('text')}",
                "",
                "ನಾನು ಎಲ್ಲವೂ ಸರಿಯಾಗಿದೆ ಎಂದು ಪರಿಶೀಲಿಸಿದ್ದೇನೆ.",
                "",
                "_" * 40,
                "Customer Signature",
                "\n\n\n"  # Extra lines for paper feed
            ])
            
            # Print the content
            with Printer(printer_name=printer_name) as printer:
                # Set printer properties for thermal paper
                printer.paper_size = 'A4'  # or whatever size your thermal paper is
                printer.orientation = 'portrait'
                
                # Print each line
                for line in lines:
                    printer.text(line + "\n")
            
            messagebox.showinfo("Success", "Invoice sent to printer!")
            
        except Exception as e:
            error_msg = f"Error printing invoice: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("Print Error", error_msg)

    def show_print_preview(self):
        """Show a preview and handle printing."""
        try:
            # Create preview window
            preview = ctk.CTkToplevel(self)
            preview.title("Print Preview")
            preview.geometry("400x600")
            preview.grab_set()  # Make the window modal
            
            # Center the preview window
            x = self.winfo_x() + (self.winfo_width() - 400) // 2
            y = self.winfo_y() + (self.winfo_height() - 600) // 2
            preview.geometry(f"400x600+{x}+{y}")

            # Create a frame for the preview content using monospace font to match printer output
            preview_frame = ctk.CTkScrollableFrame(preview)
            preview_frame.pack(fill="both", expand=True, padx=10, pady=10)

            # Preview content using monospace font to match printer output
            preview_text = ctk.CTkTextbox(preview_frame, font=("Courier New", 10))
            preview_text.pack(fill="both", expand=True)
            
            # Generate the same text content that will be printed
            lines = []
            
            # Header
            lines.extend([
                "          |ಶ್ರೀ|",
                "",
                "    G.V. Mahant Brothers",
                datetime.now().strftime("%A, %d %B %Y %I:%M %p"),
                "-" * 40,
                f"Customer: {self.customer_entry.get().strip()}",
                f"Mode: {self.current_mode.get()}",
                "-" * 40,
            ])
            
            # Add headers based on mode
            if self.current_mode.get() == "Patti":
                lines.append("Item     Pkt   Qty   Rate  Ham    Amt")
            elif self.current_mode.get() == "Kata":
                lines.append("Item     Net   Less  Rate  Ham    Amt")
            elif self.current_mode.get() == "Barthe":
                lines.append("Item     Pkt   Wt    +/-   Rate   Amt")
            
            lines.append("-" * 40)
            
            # Add rows
            for _, widgets in self.rows:
                if any(w.get().strip() for w in widgets[:-1]):
                    row_text = []
                    # Item name (limited to 8 chars)
                    row_text.append(f"{widgets[0].get()[:8]:<8}")
                    
                    # Other fields with appropriate spacing
                    for w in widgets[1:-1]:
                        row_text.append(f"{w.get()[:5]:>5}")
                    
                    # Amount
                    amount = widgets[-1].cget("text").replace("₹", "")
                    row_text.append(f"{amount:>6}")
                    
                    lines.append("".join(row_text))
            
            lines.extend([
                "-" * 40,
                f"Total Amount: {self.total_label.cget('text')}",
                "",
                "ನಾನು ಎಲ್ಲವೂ ಸರಿಯಾಗಿದೆ ಎಂದು ಪರಿಶೀಲಿಸಿದ್ದೇನೆ.",
                "",
                "_" * 40,
                "Customer Signature",
                "\n\n\n"  # Extra lines for paper feed
            ])
            
            # Insert all lines into preview
            preview_text.insert("1.0", "\n".join(lines))
            preview_text.configure(state="disabled")  # Make read-only

            # Create buttons frame
            button_frame = ctk.CTkFrame(preview, fg_color="transparent")
            button_frame.pack(side="bottom", fill="x", padx=10, pady=10)

            # Add Print and Close buttons
            ctk.CTkButton(
                button_frame,
                text="Print",
                command=lambda: [preview.destroy(), self.save_for_print()],
                width=120
            ).pack(side="left", padx=5)

            ctk.CTkButton(
                button_frame,
                text="Close",
                command=preview.destroy,
                width=120
            ).pack(side="right", padx=5)

        except Exception as e:
            error_msg = f"Preview error: {str(e)}"
            logging.error(error_msg)
            messagebox.showerror("Preview Error", error_msg)

if __name__ == "__main__":
    app = InvoiceApp()
    app.mainloop()
