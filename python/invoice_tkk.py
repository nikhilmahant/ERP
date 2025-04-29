import sys
import os
import json
import logging
from datetime import datetime
from tkinter import messagebox, ttk
import tkinter as tk
from constants import *  # Import all constants
from openpyxl import Workbook, load_workbook
import win32print
import win32api
from PIL import Image, ImageDraw, ImageFont
import io

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Define fonts for CustomTkinter
HEADER_FONT = ("Arial", 24, "bold")
SUBHEADER_FONT = ("Arial", 14)
LABEL_FONT = ("Arial", 12)
ENTRY_FONT = ("Arial", 12)
TABLE_HEADER_FONT = ("Arial", 12, "bold")
TABLE_FONT = ("Arial", 12)
BUTTON_FONT = ("Arial", 12)
LABEL_FONT_BOLD = ("Arial", 12, "bold")

MODE_HEADERS = {
    "Patti": ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount", ""],
    "Kata": ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount", ""],
    "Barthe": ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount", ""],
}

validate_float = lambda v: float(v) if v.strip() else 0 if v else 0

def safe_float(val):
    try:
        return float(val) if val.strip() else 0
    except Exception:
        return 0

class InvoiceTable(tk.Frame):
    def __init__(self, master, mode="Patti", total_label=None, **kwargs):
        super().__init__(master, **kwargs)
        self.configure(bg=FRAME_COLOR)
        self.rows = []
        self.header_widgets = []
        self.table_frame = tk.Frame(self, bg=FRAME_COLOR)
        self.table_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.mode = mode
        # If total_label is not provided, try to find it from the parent
        if total_label is None and hasattr(master, 'total_label'):
            self.total_label = master.total_label
        else:
            self.total_label = total_label
        self.master_ref = master  # Keep reference to master window
        self._draw_headers()
        self.add_row()

    def update_amounts(self):
        total = 0.0
        for row_widgets in self.rows:
            try:
                values = self.get_row_values(row_widgets)
                if not values[0].strip():
                    row_widgets[-2].configure(text="₹0.00")
                    continue
                amount = 0.0
                if self.mode == "Patti":
                    pkt = safe_float(values[1])
                    qty = safe_float(values[2])
                    rate = safe_float(values[3])
                    hamali = safe_float(values[4])
                    amount = (qty * rate) + (pkt * hamali)
                elif self.mode == "Kata":
                    net = safe_float(values[1])
                    less = safe_float(values[2])
                    final_wt = net * (1 - less / 100.0) if less < 100 else 0.0
                    rate = safe_float(values[3])
                    hamali_rate = safe_float(values[4])
                    packets = int(net / 60) if net > 0 else 0
                    amount = (final_wt * rate) + (packets * hamali_rate)
                elif self.mode == "Barthe":
                    pkt = safe_float(values[1])
                    wt_per_pkt = safe_float(values[2])
                    adj = safe_float(values[3])
                    rate = safe_float(values[4])
                    hamali_per_pkt = safe_float(values[5])
                    total_qty = (pkt * wt_per_pkt) + adj
                    amount = (total_qty * rate) + (pkt * hamali_per_pkt)
                row_widgets[-2].configure(text=f"₹{amount:.2f}")
                total += amount
            except Exception as e:
                logging.error(f"Error calculating amount for row: {e}")
        # Add kata amount if in Kata mode
        if self.mode == "Kata" and hasattr(self.master_ref, 'kata_amount_entry'):
            try:
                kata_amt = safe_float(self.master_ref.kata_amount_entry.get())
                total += kata_amt
            except Exception as e:
                logging.error(f"Error reading kata amount: {e}")
        # Update the total label if it exists
        if self.total_label:
            self.total_label.configure(text=f"Amount: ₹{total:.2f}")
        return total

    def get_row_values(self, row_widgets):
        values = []
        for i, widget in enumerate(row_widgets):
            if i == 0:  # Item combobox
                values.append(widget.get())
            elif i == len(row_widgets)-2:  # Amount label
                values.append(widget.cget("text"))
            elif i < len(row_widgets)-2:  # Other entries
                values.append(widget.get())
        return values

    def delete_row(self, row_widgets):
        for w in row_widgets:
            w.grid_forget()
            w.destroy()
        self.rows.remove(row_widgets)
        self.update_amounts()

    def clear_rows(self, draw_headers=True):
        for widgets in self.rows:
            for w in widgets:
                w.grid_forget()
                w.destroy()
        self.rows.clear()
        if draw_headers:
            self.add_row()
        self.update_amounts()


    def _draw_headers(self):
        for w in self.header_widgets:
            w.destroy()
        self.header_widgets.clear()
        for col, header in enumerate(MODE_HEADERS[self.mode]):
            lbl = tk.Label(self.table_frame, text=header, font=TABLE_HEADER_FONT, fg=PRIMARY_COLOR)
            lbl.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
            self.header_widgets.append(lbl)
            self.table_frame.grid_columnconfigure(col, weight=1)

    def set_mode(self, mode):
        self.mode = mode
        self.clear_rows(draw_headers=False)
        self._draw_headers()
        self.add_row()

    def add_row(self):
        row_idx = len(self.rows) + 1
        row_widgets = []

        # Item Combobox (ttk)
        item_combo = ttk.Combobox(self.table_frame, values=ITEM_LIST, width=22, font=("Arial", 12))
        item_combo.set("")
        item_combo.grid(row=row_idx, column=0, padx=2, pady=2, sticky="nsew")
        item_combo.bind("<KeyRelease>", lambda e: self.update_amounts())
        item_combo.bind("<<ComboboxSelected>>", lambda e: self.update_amounts())
        row_widgets.append(item_combo)

        # Entry widgets
        for col in range(1, len(MODE_HEADERS[self.mode])-2):
            entry = tk.Entry(self.table_frame, font=TABLE_FONT, width=80)
            entry.grid(row=row_idx, column=col, padx=2, pady=2, sticky="nsew")
            entry.bind("<KeyRelease>", lambda e: self.update_amounts())
            row_widgets.append(entry)

        # Amount Label
        amount_label = tk.Label(self.table_frame, text="₹0.00", font=TABLE_FONT, fg=PRIMARY_COLOR)
        amount_label.grid(row=row_idx, column=len(MODE_HEADERS[self.mode])-2, padx=2, pady=2, sticky="nsew")
        row_widgets.append(amount_label)

        # Delete Button
        del_btn = tk.Button(self.table_frame, text="🗑", width=30, bg=ERROR_COLOR, font=TABLE_FONT, command=lambda: self.delete_row(row_widgets))
        del_btn.grid(row=row_idx, column=len(MODE_HEADERS[self.mode])-1, padx=2, pady=2, sticky="nsew")
        row_widgets.append(del_btn)

        self.rows.append(row_widgets)

        self.update_amounts()

        # Focus the new Item Combobox and open dropdown safely
        item_combo.focus_set()
        def open_dropdown_safe(widget=item_combo):
            try:
                if widget.winfo_exists() and widget.winfo_ismapped():
                    widget.event_generate('<Down>')
            except Exception as e:
                logging.warning(f"Could not open combobox dropdown: {e}")
        self.after(100, open_dropdown_safe)

    def delete_row(self, row_widgets):
        for w in row_widgets:
            w.grid_forget()
            w.destroy()
        self.rows.remove(row_widgets)
        self.update_amounts()

    def clear_rows(self, draw_headers=True):
        for widgets in self.rows:
            for w in widgets:
                w.grid_forget()
                w.destroy()
        self.rows.clear()
        if draw_headers:
            self.add_row()
        self.update_amounts()

    def get_row_values(self, row_widgets):
        values = []
        for i, widget in enumerate(row_widgets):
            if i == 0:
                values.append(widget.get())
            elif i == len(row_widgets)-2:
                values.append(widget.cget("text"))
            elif i < len(row_widgets)-2:
                values.append(widget.get())
        return values

    def update_row_amounts(self):
        for row_widgets in self.rows:
            try:
                values = self.get_row_values(row_widgets)
                if not values[0].strip():
                    row_widgets[-2].configure(text="₹0.00")
                    continue
                amount = 0.0
                if self.mode == "Patti":
                    pkt = safe_float(values[1])
                    qty = safe_float(values[2])
                    rate = safe_float(values[3])
                    hamali = safe_float(values[4])
                    amount = (qty * rate) + (pkt * hamali)
                elif self.mode == "Kata":
                    net = safe_float(values[1])
                    less = safe_float(values[2])
                    final_wt = net * (1 - less / 100.0) if less < 100 else 0.0
                    rate = safe_float(values[3])
                    hamali_rate = safe_float(values[4])
                    packets = int(net / 60) if net > 0 else 0
                    amount = (final_wt * rate) + (packets * hamali_rate)
                elif self.mode == "Barthe":
                    pkt = safe_float(values[1])
                    wt_per_pkt = safe_float(values[2])
                    adj = safe_float(values[3])
                    rate = safe_float(values[4])
                    hamali_per_pkt = safe_float(values[5])
                    total_qty = (pkt * wt_per_pkt) + adj
                    amount = (total_qty * rate) + (pkt * hamali_per_pkt)
                row_widgets[-2].configure(text=f"₹{amount:.2f}")
            except Exception as e:
                logging.error(f"Error calculating row amount: {e}")

    def update_total_amount(self):
        total = 0.0
        for row_widgets in self.rows:
            amount_str = row_widgets[-2].cget("text").replace("₹", "").strip()
            total += safe_float(amount_str)

        # Add Kata amount if applicable (robust reference)
        kata_amt = 0.0
        master = getattr(self, 'master_ref', self.master)
        if self.mode == "Kata" and hasattr(master, 'kata_amount_entry') and getattr(master, 'kata_amount_entry') is not None:
            try:
                kata_amt = safe_float(master.kata_amount_entry.get())
                total += kata_amt
            except Exception as e:
                logging.error(f"Error reading kata amount: {e}")
        if self.total_label:
            self.total_label.configure(text=f"Amount: ₹{total:.2f}")

    def update_amounts(self):
        self.update_row_amounts()
        self.update_total_amount()

    def get_all_data(self):
        data = []
        for row_widgets in self.rows:
            values = self.get_row_values(row_widgets)
            if any(v.strip() for v in values):
                data.append(values)
        return data

    def delete_row(self, row_widgets):
        for w in row_widgets:
            w.grid_forget()
            w.destroy()
        self.rows.remove(row_widgets)
        self.update_amounts()

    def clear_rows(self, draw_headers=True):
        for widgets in self.rows:
            for w in widgets:
                w.grid_forget()
                w.destroy()
        self.rows.clear()
        if draw_headers:
            self.add_row()
        self.update_amounts()

    def get_row_values(self, row_widgets):
        values = []
        for i, widget in enumerate(row_widgets):
            if i == 0:
                values.append(widget.get())
            elif i == len(row_widgets)-2:
                values.append(widget.cget("text"))
            elif i < len(row_widgets)-2:
                values.append(widget.get())
        return values

    def update_amounts(self):
        total = 0.0
        for row_widgets in self.rows:
            try:
                values = self.get_row_values(row_widgets)
                if not values[0].strip():
                    row_widgets[-2].configure(text="₹0.00")
                    continue
                amount = 0.0
                if self.mode == "Patti":
                    pkt = safe_float(values[1])
                    qty = safe_float(values[2])
                    rate = safe_float(values[3])
                    hamali = safe_float(values[4])
                    amount = (qty * rate) + (pkt * hamali)
                elif self.mode == "Kata":
                    net = safe_float(values[1])
                    less = safe_float(values[2])
                    final_wt = net * (1 - less / 100.0) if less < 100 else 0.0
                    rate = safe_float(values[3])
                    hamali_rate = safe_float(values[4])
                    packets = int(net / 60) if net > 0 else 0
                    amount = (final_wt * rate) + (packets * hamali_rate)
                elif self.mode == "Barthe":
                    pkt = safe_float(values[1])
                    wt_per_pkt = safe_float(values[2])
                    adj = safe_float(values[3])
                    rate = safe_float(values[4])
                    hamali_per_pkt = safe_float(values[5])
                    total_qty = (pkt * wt_per_pkt) + adj
                    amount = (total_qty * rate) + (pkt * hamali_per_pkt)
                row_widgets[-2].configure(text=f"₹{amount:.2f}")
                total += amount
            except Exception as e:
                logging.error(f"Error calculating amount for row: {e}")
        # Fix: Use self.master.kata_amount_entry instead of self.master.master.kata_amount_entry
        if hasattr(self.master.master, 'total_label'):
            kata_amt = 0.0
            if self.mode == "Kata" and hasattr(self.master, 'kata_amount_entry') and self.master.kata_amount_entry is not None:
                try:
                    kata_amt = safe_float(self.master.kata_amount_entry.get())
                except Exception as e:
                    logging.error(f"Error reading kata_amount_entry: {e}")
            self.master.master.total_label.configure(text=f"Amount: ₹{total+kata_amt:.2f}")
        return total

    def get_all_data(self):
        data = []
        for row_widgets in self.rows:
            values = self.get_row_values(row_widgets)
            if any(v.strip() for v in values):
                data.append(values)
        return data

class PrintPreviewDialog(tk.Toplevel):
    def __init__(self, parent, content):
        super().__init__(parent)
        self.title("Print Preview")
        
        # Calculate window size based on content
        lines = content.split('\n')
        width = 400  # Fixed width for invoice
        height = min(600, len(lines) * 20 + 100)  # 20 pixels per line + padding
        
        self.geometry(f"{width}x{height}")
        self.resizable(False, False)  # Make window non-resizable
        
        # Make this window modal and always on top
        self.transient(parent)
        self.grab_set()
        
        # Create a frame for the preview content
        self.preview_frame = tk.Frame(self)
        self.preview_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create a text widget for the preview with a font that supports Kannada
        self.preview_text = tk.Text(
            self.preview_frame,
            font=("Noto Sans Kannada", 12),
            wrap="none",
            width=width - 20,  # Account for padding
            height=height - 100  # Account for button and padding
        )
        self.preview_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Add the content to the preview
        self.preview_text.insert("1.0", content)
        self.preview_text.configure(state="disabled")
        
        # Add print button
        print_btn = tk.Button(
            self,
            text="Print",
            command=self.print_invoice,
            bg=PRIMARY_COLOR,
            
            font=BUTTON_FONT
        )
        print_btn.pack(pady=10)
        
        self.content = content
        
        # Bring the window to the front
        self.lift()
        self.focus_force()

    def print_invoice(self):
        try:
            printer_name = win32print.GetDefaultPrinter()
            hPrinter = win32print.OpenPrinter(printer_name)
            
            try:
                # Initialize printer (ESC @)
                init_printer = b'\x1B\x40'
                
                # Start print job in RAW mode
                hJob = win32print.StartDocPrinter(hPrinter, 1, ("Invoice", None, "RAW"))
                win32print.StartPagePrinter(hPrinter)
                
                # Send initialization command
                win32print.WritePrinter(hPrinter, init_printer)
                
                # Print content line by line with proper encoding
                for line in self.content.split('\n'):
                    # Convert to bytes with correct encoding (try ASCII first)
                    try:
                        encoded_line = line.encode('ascii') + b'\n'
                    except UnicodeEncodeError:
                        # Fallback to UTF-8 if ASCII fails
                        encoded_line = line.encode('utf-8') + b'\n'
                    
                    win32print.WritePrinter(hPrinter, encoded_line)
                    win32api.Sleep(50)  # Small delay between lines
                
                # Paper feed and cut commands
                win32print.WritePrinter(hPrinter, b'\n\n\n\n')  # Feed 4 lines
                win32print.WritePrinter(hPrinter, b'\x1D\x56\x00')  # ESC/POS cut command
                win32api.Sleep(300)  # Wait for cut to complete
                
                win32print.EndPagePrinter(hPrinter)
                win32print.EndDocPrinter(hPrinter)
                
                messagebox.showinfo("Success", "Printed successfully with paper cut!")
                
            except Exception as e:
                logging.error(f"Printing error: {e}")
                messagebox.showerror("Print Error", f"Failed to print: {str(e)}")
            finally:
                win32print.ClosePrinter(hPrinter)
                
        except Exception as e:
            logging.error(f"Printer connection error: {e}")
            messagebox.showerror("Printer Error", f"Cannot connect to printer: {str(e)}")

class InvoiceWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        # Try all methods to maximize window on startup
        def maximize_window():
            try:
                self.state('zoomed')
            except Exception:
                pass
            try:
                self.attributes('-zoomed', True)
            except Exception:
                pass
            try:
                self.geometry(f"{self.winfo_screenwidth()}x{self.winfo_screenheight()}+0+0")
            except Exception:
                pass
        self.after(0, maximize_window)

        self.title("G.V. Mahant Brothers - Invoice")
        self.geometry("1200x800")
        self.configure(bg=BACKGROUND_COLOR)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.main_frame = tk.Frame(self, bg=BACKGROUND_COLOR)
        self.main_frame.grid(row=0, column=0, sticky="nsew")
        self.main_frame.grid_rowconfigure(4, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self._setup_header()
        self._setup_mode_selection()
        self._setup_customer_section()
        self._setup_table()
        self._setup_bottom_section()
        self.switch_mode("Patti")

    def _setup_header(self):
        # Unified header frame with three columns
        self.header_frame = tk.Frame(self.main_frame, bg=PRIMARY_COLOR)
        self.header_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
        self.header_frame.grid_columnconfigure(0, weight=1)  # left spacer
        self.header_frame.grid_columnconfigure(1, weight=2)  # center
        self.header_frame.grid_columnconfigure(2, weight=1)  # right (date)

        # Left-align company name
        company_label = tk.Label(self.header_frame, text="G.V. Mahant Brothers", font=HEADER_FONT, fg=BACKGROUND_COLOR, anchor="w", justify="left")
        company_label.grid(row=0, column=0, sticky="w", padx=(20, 0), pady=(10, 10))
        # Date label right-aligned, bold font
        self.date_label = tk.Label(self.header_frame, font=("Arial", 14, "bold"), fg=BACKGROUND_COLOR, anchor="e", justify="right")
        self.date_label.grid(row=0, column=2, sticky="e", padx=20, pady=10)

        self._update_datetime()

    def _update_datetime(self):
        now = datetime.now()
        self.date_label.configure(text=now.strftime("%A, %d %B %Y\n%I:%M %p"))
        self.after(1000, self._update_datetime)

    def _setup_mode_selection(self):
        self.mode_frame = tk.Frame(self.main_frame, bg=FRAME_COLOR)
        self.mode_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(16, 0))
        self.mode_var = tk.StringVar(value="Patti")
        # Use CTkButton widgets for mode selection styled like the Add button
        self.mode_buttons = {}
        btn_container = tk.Frame(self.mode_frame, bg=FRAME_COLOR)
        btn_container.pack(padx=10, pady=8)
        for mode in ["Patti", "Kata", "Barthe"]:
            btn = tk.Button(
                btn_container,
                text=mode,
                font=LABEL_FONT,
                bg=PRIMARY_COLOR if mode == self.mode_var.get() else ACCENT_COLOR,
                fg=BACKGROUND_COLOR if mode == self.mode_var.get() else TEXT_COLOR,
                
                
                width=90,
                command=lambda m=mode: self._on_mode_button_click(m)
            )
            btn.pack(side="left", padx=8)
            self.mode_buttons[mode] = btn

    def _on_mode_button_click(self, mode):
        self.mode_var.set(mode)
        self.switch_mode(mode)
        # Update button styles
        for m, btn in self.mode_buttons.items():
            if m == mode:
                btn.configure(bg=PRIMARY_COLOR, fg=BACKGROUND_COLOR)
            else:
                btn.configure(bg=ACCENT_COLOR, fg=TEXT_COLOR)

    def _on_mode_change(self):
        self.switch_mode(self.mode_var.get())

    def switch_mode(self, mode):
        self.current_mode = mode
        self.table.set_mode(mode)
        self._setup_kata_field()
        self.table.update_amounts()

    def _setup_customer_section(self):
        self.customer_frame = tk.Frame(self.main_frame, bg=BACKGROUND_COLOR)
        self.customer_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=(10, 0))
        tk.Label(self.customer_frame, text="Customer Name:", font=LABEL_FONT, fg=TEXT_COLOR).pack(side="left", padx=10)
        self.customer_entry = tk.Entry(self.customer_frame, font=ENTRY_FONT, width=400)
        self.customer_entry.pack(side="left", padx=10)
        tk.Label(self.customer_frame, text="", bg=BACKGROUND_COLOR).pack(side="left", expand=True, fill="x")

    def _setup_table(self):
        # Create a frame to hold the canvas and scrollbar
        self.table_outer_frame = tk.Frame(self.main_frame, bg=BACKGROUND_COLOR)
        self.table_outer_frame.grid(row=4, column=0, sticky="nsew", padx=10, pady=(10, 0))
        self.main_frame.grid_rowconfigure(4, weight=1)
        self.table_outer_frame.grid_rowconfigure(0, weight=1)
        self.table_outer_frame.grid_columnconfigure(0, weight=1)

        # Create canvas and scrollbar
        self.table_canvas = tk.Canvas(self.table_outer_frame, bg=BACKGROUND_COLOR, highlightthickness=0)
        self.table_canvas.grid(row=0, column=0, sticky="nsew")
        self.table_scrollbar = ttk.Scrollbar(self.table_outer_frame, orient="vertical", command=self.table_canvas.yview)
        # Don't grid the scrollbar yet; it will be shown only when needed
        self.table_canvas.configure(yscrollcommand=self.table_scrollbar.set)

        # Create a frame inside the canvas for the InvoiceTable
        self.table_inner_frame = tk.Frame(self.table_canvas, bg=BACKGROUND_COLOR)
        self.table_window = self.table_canvas.create_window((0, 0), window=self.table_inner_frame, anchor="nw")

        # Add the InvoiceTable to the inner frame
        # Pass the total_label reference from the main window to the table
        self.table = InvoiceTable(self.table_inner_frame, mode="Patti", total_label=self.master.total_label if hasattr(self.master, 'total_label') else None)
        self.table.pack(fill="both", expand=True)

        # Scroll region update on resize and show/hide scrollbar
        def _on_frame_configure(event):
            self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all"))
            # Show scrollbar only if needed
            canvas_height = self.table_canvas.winfo_height()
            frame_height = self.table_inner_frame.winfo_height()
            if frame_height > canvas_height:
                self.table_scrollbar.grid(row=0, column=1, sticky="ns")
            else:
                self.table_scrollbar.grid_remove()
        self.table_inner_frame.bind("<Configure>", _on_frame_configure)

        # Mousewheel scrolling only when mouse is over the canvas and scrollbar is visible
        def _on_mousewheel(event):
            # Only scroll if scrollbar is visible
            if self.table_scrollbar.winfo_ismapped():
                self.table_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        def _bind_mousewheel(event):
            self.table_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        def _unbind_mousewheel(event):
            self.table_canvas.unbind_all("<MouseWheel>")
        self.table_canvas.bind("<Enter>", _bind_mousewheel)
        self.table_canvas.bind("<Leave>", _unbind_mousewheel)

        # Ensure canvas expands
        self.table_canvas.bind("<Configure>", lambda e: self.table_canvas.itemconfig(self.table_window, width=e.width))


    def _setup_bottom_section(self):
        self.bottom_frame = tk.Frame(self.main_frame, bg=BACKGROUND_COLOR)
        self.bottom_frame.grid(row=5, column=0, sticky="ew", padx=10, pady=10)
        self.bottom_frame.grid_columnconfigure(0, weight=0)  # buttons
        self.bottom_frame.grid_columnconfigure(1, weight=1)  # spacer
        self.bottom_frame.grid_columnconfigure(2, weight=0)  # kata/total

        # Kata and total (right, now placed above the button group)
        self.right_bottom_frame = tk.Frame(self.bottom_frame, bg=BACKGROUND_COLOR)
        self.right_bottom_frame.grid(row=0, column=0, columnspan=3, sticky="ew")
        self.right_bottom_frame.grid_columnconfigure(0, weight=1)
        self.kata_frame = tk.Frame(self.right_bottom_frame, bg=BACKGROUND_COLOR)
        self.kata_frame.grid(row=0, column=0, sticky="e", padx=(0, 10))
        self.total_label = tk.Label(
            self.right_bottom_frame,
            text="Amount: ₹0.00",
            fg=PRIMARY_COLOR,
            font=HEADER_FONT
        )
        self.total_label.grid(row=0, column=1, sticky="e", pady=(2,0))

        # Button group (left)
        btn_frame = tk.Frame(self.bottom_frame, bg=BACKGROUND_COLOR)
        btn_frame.grid(row=1, column=0, sticky="w")
        add_row_btn = tk.Button(
            btn_frame,
            text="Add Row",
            command=self.table.add_row,
            bg=PRIMARY_COLOR,
            
            font=BUTTON_FONT
        )
        add_row_btn.pack(side="left", padx=5)
        clear_btn = tk.Button(
            btn_frame,
            text="Clear",
            command=self.table.clear_rows,
            bg=PRIMARY_COLOR,
            
            font=BUTTON_FONT
        )
        clear_btn.pack(side="left", padx=5)
        save_btn = tk.Button(
            btn_frame,
            text="Save",
            command=self.save_to_excel,
            bg=PRIMARY_COLOR,
            
            font=BUTTON_FONT
        )
        save_btn.pack(side="left", padx=5)
        print_btn = tk.Button(
            btn_frame,
            text="Print",
            command=self.show_print_preview,
            bg=PRIMARY_COLOR,
            
            font=BUTTON_FONT
        )
        print_btn.pack(side="left", padx=5)

        # Spacer (column 1) is empty for stretch


    def _setup_kata_field(self):
        for w in self.kata_frame.winfo_children():
            w.destroy()
        if self.current_mode == "Kata":
            # Place kata label and entry horizontally using grid for alignment
            kata_label = tk.Label(self.kata_frame, text="Kata:", font=LABEL_FONT_BOLD, fg=TEXT_COLOR)
            kata_label.grid(row=0, column=0, padx=(0, 8), pady=(0, 0), sticky="e")
            self.kata_amount_entry = tk.Entry(self.kata_frame, font=ENTRY_FONT, width=120)
            self.kata_amount_entry.insert(0, "0")
            self.kata_amount_entry.grid(row=0, column=1, padx=(0, 10), pady=(0, 0), sticky="e")
            self.kata_amount_entry.bind("<KeyRelease>", lambda e: self.table.update_amounts())
            self.kata_amount_entry.bind("<FocusOut>", lambda e: self.table.update_amounts())
            # Ensure total updates after kata field is created
            self.after(50, self.table.update_amounts)
        else:
            self.kata_amount_entry = None
            self.table.update_amounts()

    def save_to_excel(self):
        try:
            home_dir = os.path.expanduser("~")
            documents_path = os.path.join(home_dir, "Documents")
            os.makedirs(documents_path, exist_ok=True)
            date_str = datetime.now().strftime('%Y-%m-%d')
            base_filename = f"Invoice_{date_str}.xlsx"
            full_save_path = os.path.join(documents_path, base_filename)
            customer = self.customer_entry.get().strip() or "Unknown Customer"
            mode = self.current_mode
            headers = MODE_HEADERS[mode]
            data_rows = self.table.get_all_data()
            if not data_rows:
                messagebox.showwarning("No Data", "No data entered to save.")
                return
            try:
                if os.path.exists(full_save_path):
                    wb = load_workbook(full_save_path)
                else:
                    wb = Workbook()
                if mode in wb.sheetnames:
                    ws = wb[mode]
                else:
                    if len(wb.sheetnames) > 0:
                        ws = wb.create_sheet(title=mode)
                    else:
                        ws = wb.active
                        ws.title = mode
                ws.delete_rows(1, ws.max_row)
                ws.append(["Timestamp", "Customer"] + headers)
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                for row in data_rows:
                    ws.append([timestamp, customer] + row)
                wb.save(full_save_path)
                messagebox.showinfo("Saved", f"Invoice data saved to:\n{full_save_path}\n(Sheet: {mode})")
            except PermissionError:
                messagebox.showerror("Permission Error", f"Cannot save '{base_filename}'.\nThe file might be open in Excel.\n\nLocation: {documents_path}")
            except Exception as e:
                messagebox.showerror("Save Error", f"Error saving Excel file to:\n{full_save_path}\n\nError: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error during save operation: {str(e)}")

    def show_print_preview(self):
        """Show a preview of the invoice before printing."""
        try:
            customer_name = self.customer_entry.get().strip() or "Unknown Customer"
            data_rows = self.table.get_all_data()
            
            if not data_rows:
                messagebox.showwarning("No Data", "No data to print.")
                return

            # Convert data rows to invoice items format
            invoice_items = []
            for row in data_rows:
                if self.current_mode == "Patti":
                    item = {
                        "item": row[0],
                        "packet": float(row[1] or 0),
                        "quantity": float(row[2] or 0),
                        "rate": float(row[3] or 0),
                        "hamali": float(row[4] or 0)
                    }
                elif self.current_mode == "Kata":
                    item = {
                        "item": row[0],
                        "net_wt": float(row[1] or 0),
                        "less_percent": float(row[2] or 0),
                        "rate": float(row[3] or 0),
                        "hamali_rate": float(row[4] or 0)
                    }
                elif self.current_mode == "Barthe":
                    item = {
                        "item": row[0],
                        "packet": float(row[1] or 0),
                        "weight": float(row[2] or 0),
                        "adjustment": float(row[3] or 0),
                        "rate": float(row[4] or 0),
                        "hamali": float(row[5] or 0)
                    }
                invoice_items.append(item)

            # Prepare the content with exact formatting
            content = []
            content.append("          G.V. Mahant Brothers          ")
            content.append(datetime.now().strftime("          %d-%b-%Y %I:%M %p          "))
            content.append("-" * 48)
            content.append(f"Customer: {customer_name}")
            content.append("-" * 48)

            # Add appropriate headers based on mode
            if self.current_mode == "Patti":
                content.append(f"{'Item':<10}    {'Pkt':>3}  {'Qty':>3}  {'Rate':>5}   {'Ham':>3}      {'Amt':>5}")
            elif self.current_mode == "Kata":
                content.append(f"{'Item':<10}    {'Net':>3}  {'Less%':>5}  {'Rate':>5}   {'Ham':>3}      {'Amt':>5}")
            elif self.current_mode == "Barthe":
                content.append(f"{'Item':<10}    {'Pkt':>3}  {'Wt':>3}  {'Adj':>3}  {'Rate':>5}   {'Ham':>3}      {'Amt':>5}")
            
            content.append("-" * 48)

            # Add items
            total_amount = 0
            for item in invoice_items:
                if self.current_mode == "Patti":
                    item_name = item['item'][:10]
                    packet = item['packet']
                    quantity = item['quantity']
                    rate = item['rate']
                    hamali = item['hamali']
                    amount = (quantity * rate) + (packet * hamali)
                    content.append(f"{item_name:<10}    {packet:>3.1f}  {quantity:>3.1f}  {rate:>5.1f}   {hamali:>3.1f}    {amount:>7.2f}")
                
                elif self.current_mode == "Kata":
                    item_name = item['item'][:10]
                    net_wt = item['net_wt']
                    less_percent = item['less_percent']
                    rate = item['rate']
                    hamali_rate = item['hamali_rate']
                    final_wt = net_wt * (1 - less_percent / 100.0) if less_percent < 100 else 0.0
                    packets = int(net_wt / 60) if net_wt > 0 else 0
                    amount = (final_wt * rate) + (packets * hamali_rate)
                    content.append(f"{item_name:<10}    {net_wt:>3.1f}  {less_percent:>5.1f}  {rate:>5.1f}   {hamali_rate:>3.1f}    {amount:>7.2f}")
                
                elif self.current_mode == "Barthe":
                    item_name = item['item'][:10]
                    packet = item['packet']
                    weight = item['weight']
                    adjustment = item['adjustment']
                    rate = item['rate']
                    hamali = item['hamali']
                    total_qty = (packet * weight) + adjustment
                    amount = (total_qty * rate) + (packet * hamali)
                    content.append(f"{item_name:<10}    {packet:>3.1f}  {weight:>3.1f}  {adjustment:>3.1f}  {rate:>5.1f}   {hamali:>3.1f}    {amount:>7.2f}")
                
                total_amount += amount

            # Add kata amount if in kata mode
            if self.current_mode == "Kata" and hasattr(self, 'kata_amount_entry'):
                kata_amount = float(self.kata_amount_entry.get() or 0)
                total_amount += kata_amount
                content.append("-" * 48)
                content.append(f"{'':>14}Kata Amount: {kata_amount:>7.2f}")

            content.append("-" * 48)
            content.append(f"{'':>14}Total Amount: {total_amount:>7.2f}")
            content.append("-" * 48)
            content.append("ನಾನು ಎಲ್ಲವೂ ಸರಿಯಾಗಿದೆ ಎಂದು ಪರಿಶೀಲಿಸಿದ್ದೇನೆ.")
            content.append("")  # Empty line
            content.append("")  # Empty line
            content.append("")  # Empty line

            # Join the content with newlines
            preview_content = '\n'.join(content)
            
            # Show the preview dialog
            PrintPreviewDialog(self, preview_content)
            
        except Exception as e:
            logging.error(f"Error showing print preview: {e}")
            messagebox.showerror("Preview Error", f"Error showing print preview: {str(e)}")

if __name__ == "__main__":
    app = InvoiceWindow()
    app.mainloop()