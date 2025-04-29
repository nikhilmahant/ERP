import sys
import os
import json
import logging
from datetime import datetime
import customtkinter as ctk
from tkinter import messagebox, ttk
import tkinter as tk
from openpyxl import Workbook, load_workbook
import win32print
from PIL import Image, ImageDraw, ImageFont
import numpy as np
import subprocess
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib import colors

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filename='invoice_print.log'
)

# Constants
PRIMARY_COLOR = "#2B2D42"
SECONDARY_COLOR = "#8D99AE"
BACKGROUND_COLOR = "#EDF2F4"
FRAME_COLOR = "#FFFFFF"
TEXT_COLOR = "#2B2D42"
BORDER_COLOR = "#8D99AE"
ACCENT_COLOR = "#D90429"
ERROR_COLOR = "#EF233C"

# Fonts
HEADER_FONT = ("Arial", 24, "bold")
SUBHEADER_FONT = ("Arial", 14)
LABEL_FONT = ("Arial", 12)
ENTRY_FONT = ("Arial", 12)
TABLE_HEADER_FONT = ("Arial", 12, "bold")
TABLE_FONT = ("Arial", 12)
BUTTON_FONT = ("Arial", 12)
LABEL_FONT_BOLD = ("Arial", 12, "bold")

# Mode Headers
MODE_HEADERS = {
    "Patti": ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount", ""],
    "Kata": ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount", ""],
    "Barthe": ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount", ""],
}

# Item List
ITEM_LIST = [
    "Rice",
    "Wheat",
    "Sugar",
    "Dal",
    "Oil",
]

# Utility functions
validate_float = lambda v: float(v) if v.strip() else 0 if v else 0

def safe_float(val):
    try:
        return float(val) if val.strip() else 0
    except Exception:
        return 0

class InvoiceTable(ctk.CTkFrame):
    def __init__(self, master, mode="Patti", **kwargs):
        super().__init__(master, **kwargs)
        self.configure(fg_color=FRAME_COLOR)
        self.rows = []
        self.header_widgets = []
        self.table_frame = ctk.CTkFrame(self, fg_color=FRAME_COLOR)
        self.table_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.mode = mode
        self._draw_headers()
        self.add_row()

    def _draw_headers(self):
        for w in self.header_widgets:
            w.destroy()
        self.header_widgets.clear()
        for col, header in enumerate(MODE_HEADERS[self.mode]):
            lbl = ctk.CTkLabel(self.table_frame, text=header, font=TABLE_HEADER_FONT, text_color=PRIMARY_COLOR)
            lbl.grid(row=0, column=col, padx=2, pady=2, sticky="nsew")
            self.header_widgets.append(lbl)
            self.table_frame.grid_columnconfigure(col, weight=1)

    def set_mode(self, mode):
        self.mode = mode
        self.clear_rows(draw_headers=False)
        self._draw_headers()
        self.add_row()

    def add_row(self):
        row_idx = len(self.rows) + 1
        row_widgets = []
        item_combo = ttk.Combobox(self.table_frame, values=ITEM_LIST, width=22, font=("Arial", 12))
        item_combo.set("")
        item_combo.grid(row=row_idx, column=0, padx=2, pady=2, sticky="nsew")
        item_combo.bind("<KeyRelease>", lambda e: self.update_amounts())
        item_combo.bind("<<ComboboxSelected>>", lambda e: self.update_amounts())
        row_widgets.append(item_combo)
        for col in range(1, len(MODE_HEADERS[self.mode])-2):
            entry = ctk.CTkEntry(self.table_frame, font=TABLE_FONT, width=80)
            entry.grid(row=row_idx, column=col, padx=2, pady=2, sticky="nsew")
            entry.bind("<KeyRelease>", lambda e: self.update_amounts())
            row_widgets.append(entry)
        amount_label = ctk.CTkLabel(self.table_frame, text="₹0.00", font=TABLE_FONT, text_color=PRIMARY_COLOR)
        amount_label.grid(row=row_idx, column=len(MODE_HEADERS[self.mode])-2, padx=2, pady=2, sticky="nsew")
        row_widgets.append(amount_label)
        del_btn = ctk.CTkButton(self.table_frame, text="🗑", width=30, fg_color=ERROR_COLOR, hover_color=ACCENT_COLOR, font=TABLE_FONT, command=lambda: self.delete_row(row_widgets))
        del_btn.grid(row=row_idx, column=len(MODE_HEADERS[self.mode])-1, padx=2, pady=2, sticky="nsew")
        row_widgets.append(del_btn)
        self.rows.append(row_widgets)
        self.update_amounts()

    def delete_row(self, row_widgets):
        for w in row_widgets:
            w.grid_forget()
            w.destroy()
        self.rows.remove(row_widgets)
        self.update_amounts()

    def clear_rows(self, draw_headers=True):
        for widgets in self.rows:
            for w in widgets:
                w.grid_forget()
                w.destroy()
        self.rows.clear()
        if draw_headers:
            self.add_row()
        self.update_amounts()

    def get_row_values(self, row_widgets):
        values = []
        for i, widget in enumerate(row_widgets):
            if i == 0:
                values.append(widget.get())
            elif i == len(row_widgets)-2:
                values.append(widget.cget("text"))
            elif i < len(row_widgets)-2:
                values.append(widget.get())
        return values

    def update_amounts(self):
        total = 0.0
        for row_widgets in self.rows:
            try:
                values = self.get_row_values(row_widgets)
                if not values[0].strip():
                    row_widgets[-2].configure(text="₹0.00")
                    continue
                amount = 0.0
                if self.mode == "Patti":
                    pkt = safe_float(values[1])
                    qty = safe_float(values[2])
                    rate = safe_float(values[3])
                    hamali = safe_float(values[4])
                    amount = (qty * rate) + (pkt * hamali)
                elif self.mode == "Kata":
                    net = safe_float(values[1])
                    less = safe_float(values[2])
                    final_wt = net * (1 - less / 100.0) if less < 100 else 0.0
                    rate = safe_float(values[3])
                    hamali_rate = safe_float(values[4])
                    packets = int(net / 60) if net > 0 else 0
                    amount = (final_wt * rate) + (packets * hamali_rate)
                elif self.mode == "Barthe":
                    pkt = safe_float(values[1])
                    wt_per_pkt = safe_float(values[2])
                    adj = safe_float(values[3])
                    rate = safe_float(values[4])
                    hamali_per_pkt = safe_float(values[5])
                    total_qty = (pkt * wt_per_pkt) + adj
                    amount = (total_qty * rate) + (pkt * hamali_per_pkt)
                row_widgets[-2].configure(text=f"₹{amount:.2f}")
                total += amount
            except Exception as e:
                logging.error(f"Error calculating amount for row: {e}")
        if hasattr(self.master.master, 'total_label'):
            if self.mode == "Kata" and hasattr(self.master.master, 'kata_amount_entry') and self.master.master.kata_amount_entry is not None:
                kata_amt = safe_float(self.master.master.kata_amount_entry.get())
            else:
                kata_amt = 0.0
            self.master.master.total_label.configure(text=f"Amount: ₹{total+kata_amt:.2f}")
        return total

    def get_all_data(self):
        data = []
        for row_widgets in self.rows:
            values = self.get_row_values(row_widgets)
            if any(v.strip() for v in values):
                data.append(values)
        return data

class PrintPreviewDialog(ctk.CTkToplevel):
    def __init__(self, parent, content, mode, dynamic_content):
        super().__init__(parent)
        self.title("Print Preview")
        
        # Calculate window size based on content
        lines = content.split('\n')
        width = 400
        height = min(600, len(lines) * 20 + 100)
        
        self.geometry(f"{width}x{height}")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        
        # Create a frame for the preview content
        self.preview_frame = ctk.CTkFrame(self)
        self.preview_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Create a text widget for the preview
        self.preview_text = ctk.CTkTextbox(
            self.preview_frame,
            font=("Courier New", 12),
            wrap="none",
            width=width - 20,
            height=height - 100
        )
        self.preview_text.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Add the content to the preview
        self.preview_text.insert("1.0", content)
        self.preview_text.configure(state="disabled")
        
        # Buttons frame
        button_frame = ctk.CTkFrame(self)
        button_frame.pack(fill="x", padx=10, pady=10)
        
        # Print options button
        print_options_btn = ctk.CTkButton(
            button_frame,
            text="Print Options",
            command=self.show_print_options,
            fg_color=PRIMARY_COLOR,
            hover_color=SECONDARY_COLOR,
            font=BUTTON_FONT
        )
        print_options_btn.pack(side="left", padx=5)
        
        # Close button
        close_btn = ctk.CTkButton(
            button_frame,
            text="Close",
            command=self.destroy,
            fg_color=ERROR_COLOR,
            hover_color=ACCENT_COLOR,
            font=BUTTON_FONT
        )
        close_btn.pack(side="right", padx=5)
        
        self.content = content
        self.mode = mode
        self.dynamic_content = dynamic_content
        self.lift()
        self.focus_force()

    def show_print_options(self):
        """Show dialog for selecting print method."""
        dialog = ctk.CTkToplevel(self)
        dialog.title("Print Options")
        dialog.geometry("300x200")
        dialog.transient(self)
        dialog.grab_set()
        
        label = ctk.CTkLabel(dialog, text="Select printing method:", font=LABEL_FONT)
        label.pack(pady=20)
        
        thermal_btn = ctk.CTkButton(
            dialog,
            text="Thermal Printer",
            command=lambda: [dialog.destroy(), self.print_invoice()],
            font=BUTTON_FONT
        )
        thermal_btn.pack(pady=10)
        
        pdf_btn = ctk.CTkButton(
            dialog,
            text="Standard Printer (PDF)",
            command=lambda: [dialog.destroy(), self.print_as_pdf()],
            font=BUTTON_FONT
        )
        pdf_btn.pack(pady=10)
        
        cancel_btn = ctk.CTkButton(
            dialog,
            text="Cancel",
            command=dialog.destroy,
            fg_color=ERROR_COLOR,
            hover_color=ACCENT_COLOR,
            font=BUTTON_FONT
        )
        cancel_btn.pack(pady=10)

    def print_as_pdf(self):
        """Print invoice using ReportLab PDF generation."""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import A4
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import Paragraph, Table, TableStyle
            from reportlab.lib import colors
            import tempfile
            import os
            import win32api
            import win32print
            import win32con
            import subprocess
            
            # Check printer status
            printer_name = win32print.GetDefaultPrinter()
            logging.info(f"Using printer: {printer_name}")
            
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                printer_info = win32print.GetPrinter(hPrinter, 2)
                status = printer_info['Status']
                
                if status != 0:
                    error_msg = "Printer is not ready. Please check:\n"
                    if status & win32print.PRINTER_STATUS_OFFLINE:
                        error_msg += "- Printer is offline\n"
                    if status & win32print.PRINTER_STATUS_ERROR:
                        error_msg += "- Printer has an error\n"
                    if status & win32print.PRINTER_STATUS_NO_TONER:
                        error_msg += "- Printer is out of toner\n"
                    if status & win32print.PRINTER_STATUS_PAPER_OUT:
                        error_msg += "- Printer is out of paper\n"
                    if status & win32print.PRINTER_STATUS_PAPER_JAM:
                        error_msg += "- Printer has a paper jam\n"
                    if status & win32print.PRINTER_STATUS_DOOR_OPEN:
                        error_msg += "- Printer door is open\n"
                    
                    messagebox.showerror("Printer Error", error_msg)
                    return
            except Exception as e:
                logging.error(f"Error checking printer status: {e}")
                messagebox.showerror("Printer Error", "Could not check printer status. Please ensure the printer is properly connected and turned on.")
                return
            finally:
                win32print.ClosePrinter(hPrinter)
            
            # Create a temporary file for the PDF
            temp_handle, temp_path = tempfile.mkstemp(suffix='.pdf')
            os.close(temp_handle)
            logging.info(f"Created temporary PDF file: {temp_path}")
            
            # Register Noto Sans font for Kannada
            font_paths = [
                os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "NotoSansKannada-Regular.ttf"),
                os.path.join(os.path.dirname(__file__), "NotoSansKannada-Regular.ttf")
            ]
            
            font_path = None
            for path in font_paths:
                if os.path.exists(path):
                    font_path = path
                    break
            
            if font_path:
                pdfmetrics.registerFont(TTFont('NotoSansKannada', font_path))
                font_name = 'NotoSansKannada'
                logging.info(f"Using font: {font_path}")
            else:
                font_name = 'Helvetica'
                logging.warning("Noto Sans Kannada font not found. Using Helvetica.")
            
            # Create styles for different text elements
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(
                name='Kannada',
                fontName='NotoSansKannada',
                fontSize=12,
                leading=14,
                alignment=1  # Center alignment
            ))
            styles.add(ParagraphStyle(
                name='KannadaHeader',
                fontName='NotoSansKannada',
                fontSize=18,
                leading=20,
                alignment=1  # Center alignment
            ))
            styles.add(ParagraphStyle(
                name='KannadaTable',
                fontName='NotoSansKannada',
                fontSize=10,
                leading=12,
                alignment=0  # Left alignment
            ))
            
            # Create the PDF with proper margins and layout
            c = canvas.Canvas(temp_path, pagesize=A4)
            width, height = A4
            
            # Set margins
            left_margin = 50
            right_margin = width - 50
            top_margin = height - 50
            bottom_margin = 50
            
            # Draw header
            y = top_margin
            header_text = "ಶ್ರೀ"
            p = Paragraph(header_text, styles['KannadaHeader'])
            p.wrapOn(c, width - 2*left_margin, height)
            p.drawOn(c, left_margin, y - 20)
            y -= 40
            
            # Company name
            company_text = "G.V. Mahant Brothers"
            p = Paragraph(company_text, styles['KannadaHeader'])
            p.wrapOn(c, width - 2*left_margin, height)
            p.drawOn(c, left_margin, y - 20)
            y -= 40
            
            # Invoice text
            invoice_text = "Invoice"
            p = Paragraph(invoice_text, styles['KannadaHeader'])
            p.wrapOn(c, width - 2*left_margin, height)
            p.drawOn(c, left_margin, y - 20)
            y -= 40
            
            # Draw horizontal line
            c.line(left_margin, y, right_margin, y)
            y -= 20
            
            # Process content
            content_lines = self.content.split('\n')
            table_data = []
            current_table = []
            
            for line in content_lines:
                if line.strip():
                    if "|" in line:
                        # Table row
                        cells = [cell.strip() for cell in line.split("|") if cell.strip()]
                        current_table.append(cells)
                    else:
                        # Regular text
                        if current_table:
                            # Draw the accumulated table
                            t = Table(current_table, colWidths=[100] * len(current_table[0]))
                            t.setStyle(TableStyle([
                                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                                ('FONT', (0, 0), (-1, -1), font_name, 10),
                                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                            ]))
                            t.wrapOn(c, width - 2*left_margin, height)
                            t.drawOn(c, left_margin, y - t._height)
                            y -= t._height + 20
                            current_table = []
                        
                        # Draw regular text
                        p = Paragraph(line, styles['Kannada'])
                        p.wrapOn(c, width - 2*left_margin, height)
                        p.drawOn(c, left_margin, y - p.height)
                        y -= p.height + 10
            
            # Draw any remaining table
            if current_table:
                t = Table(current_table, colWidths=[100] * len(current_table[0]))
                t.setStyle(TableStyle([
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONT', (0, 0), (-1, -1), font_name, 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ]))
                t.wrapOn(c, width - 2*left_margin, height)
                t.drawOn(c, left_margin, y - t._height)
                y -= t._height + 20
            
            # Draw footer
            y = bottom_margin + 20
            footer_text = "ನಾನು ಎಲ್ಲವೂ ಸರಿಯಾಗಿದೆ ಎಂದು ಪರಿಶೀಲಿಸಿದ್ದೇನೆ."
            p = Paragraph(footer_text, styles['Kannada'])
            p.wrapOn(c, width - 2*left_margin, height)
            p.drawOn(c, left_margin, y)
            
            # Save the PDF
            c.save()
            logging.info("PDF saved successfully")
            
            # Try to print the PDF using different methods
            try:
                # Method 1: Using Windows print command
                print_command = f'rundll32 mshtml.dll,PrintHTML "{temp_path}"'
                result = subprocess.run(print_command, shell=True, capture_output=True, text=True)
                if result.returncode != 0:
                    logging.warning(f"Print command failed: {result.stderr}")
                    
                    # Method 2: Using win32print
                    hPrinter = win32print.OpenPrinter(printer_name)
                    try:
                        hJob = win32print.StartDocPrinter(hPrinter, 1, ("Invoice", None, "RAW"))
                        win32print.StartPagePrinter(hPrinter)
                        
                        # Read PDF file and send to printer
                        with open(temp_path, 'rb') as f:
                            win32print.WritePrinter(hPrinter, f.read())
                        
                        win32print.EndPagePrinter(hPrinter)
                        win32print.EndDocPrinter(hPrinter)
                        logging.info("Successfully sent to printer using win32print")
                    finally:
                        win32print.ClosePrinter(hPrinter)
                
                messagebox.showinfo("Success", "Invoice sent to printer!")
                self.destroy()
            except Exception as e:
                logging.error(f"Error sending to printer: {e}")
                messagebox.showerror("Print Error", f"Failed to send to printer: {str(e)}\n\nPlease check:\n1. Printer is turned on\n2. Printer is properly connected\n3. Printer has paper\n4. Printer is not in error state")
            
            # Clean up the temporary file after a delay
            def cleanup_temp_file():
                try:
                    os.unlink(temp_path)
                    logging.info("Temporary file cleaned up")
                except Exception as e:
                    logging.error(f"Error deleting temporary file: {e}")
            
            self.after(10000, cleanup_temp_file)  # Clean up after 10 seconds
            
        except Exception as e:
            logging.error(f"PDF printing error: {e}")
            messagebox.showerror("Print Error", f"Failed to print: {str(e)}\n\nPlease check:\n1. Printer is turned on\n2. Printer is properly connected\n3. Printer has paper\n4. Printer is not in error state")

    def print_invoice(self):
        try:
            # Render invoice using template and dynamic content
            self.print_as_image()
            messagebox.showinfo("Success", "Invoice printed successfully!")
            self.destroy()
        except Exception as e:
            logging.error(f"Printing error: {e}")
            messagebox.showerror("Print Error", f"Failed to print: {str(e)}")

    def print_as_image(self):
        """Render invoice using template image and overlay dynamic content."""
        # Image settings
        font_size = 18
        line_height = font_size + 5
        max_chars = 48
        width = 576  # 80mm paper at 203 DPI (adjust to 384 for 58mm)
        
        # Load template image
        try:
            image = Image.open("template.bmp").convert('L')  # Convert to grayscale
            logging.info("Successfully loaded template.bmp")
            # Save debug template
            image.save("debug_template.bmp", "BMP")
        except Exception as e:
            logging.error(f"Error loading template.bmp: {e}")
            raise Exception("Template image (template.bmp) not found. Please run create_template.py.")
        
        # Create a copy of the image for drawing
        draw_image = image.copy()
        draw = ImageDraw.Draw(draw_image)
        
        # Load Noto Sans font for Kannada
        font_paths = [
            os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "NotoSans-Regular.ttf"),
            os.path.join(os.environ.get("WINDIR", "C:\\Windows"), "Fonts", "NotoSansKannada-Regular.ttf"),
            os.path.join(os.path.dirname(__file__), "NotoSans-Regular.ttf"),
            os.path.join(os.path.dirname(__file__), "NotoSansKannada-Regular.ttf")
        ]
        kannada_font = None
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    kannada_font = ImageFont.truetype(font_path, font_size)
                    # Test if the font supports Kannada
                    if test_font_support(kannada_font, "ಶ್ರೀ"):
                        logging.info(f"Successfully loaded Noto Sans font: {font_path}")
                        break
                    else:
                        logging.warning(f"Font {font_path} lacks support for Kannada")
                        kannada_font = None
                except Exception as e:
                    logging.warning(f"Failed to load font {font_path}: {e}")
        
        if kannada_font is None:
            kannada_font = ImageFont.load_default()
            logging.warning("Noto Sans font not found or lacks support for Kannada. Using default font.")
            messagebox.showwarning("Font Warning", "Noto Sans font not found. Using default font, which may cause rendering issues.")
        
        # Start drawing dynamic content after header (line 6)
        y = 10 + 5 * line_height  # Skip 5 lines (header)
        
        # Draw dynamic content
        for idx, line in enumerate(self.dynamic_content):
            try:
                logging.info(f"Drawing dynamic line {idx}: '{line}' at y={y}")
                draw.text((10, y), line, font=kannada_font, fill=0)
                bbox = draw.textbbox((10, y), line, font=kannada_font)
                draw.rectangle(bbox, outline=128)
            except Exception as e:
                logging.error(f"Error drawing dynamic line {idx} '{line}': {e}")
            y += line_height
        
        # --- DEBUG: Draw test Kannada and English text at a fixed position ---
        test_y = 100
        try:
            # Kannada test
            test_kannada = "ಶ್ರೀ (Kannada Test)"
            draw.text((10, test_y), test_kannada, font=kannada_font, fill=0)
            bbox_k = draw.textbbox((10, test_y), test_kannada, font=kannada_font)
            draw.rectangle(bbox_k, outline=128)
            # English test
            test_english = "TEST ENGLISH"
            draw.text((10, test_y + 40), test_english, font=kannada_font, fill=0)
            bbox_e = draw.textbbox((10, test_y + 40), test_english, font=kannada_font)
            draw.rectangle(bbox_e, outline=128)
            logging.info("DREW TEST KANNADA AND ENGLISH TEXT")
        except Exception as e:
            logging.error(f"Error drawing test strings: {e}")
        # --- END DEBUG ---
        
        # Save debug image with text
        try:
            draw_image.save("debug_with_text.bmp", "BMP")
            logging.info("Debug image with text saved as debug_with_text.bmp")
        except Exception as e:
            logging.error(f"Error saving debug image with text: {e}")
        
        # Convert to 1-bit for printing
        final_image = draw_image.convert('1')
        
        # Save final image before printing
        try:
            final_image.save("debug_final.bmp", "BMP")
            logging.info("Final image saved as debug_final.bmp")
        except Exception as e:
            logging.error(f"Error saving final image: {e}")
        
        # Convert image to ESC/POS raster bitmap
        pixels = final_image.load()
        width, height = final_image.size
        
        # ESC/POS commands
        ESC = b'\x1B'
        GS = b'\x1D'
        init_cmd = ESC + b'@'  # Initialize printer
        align_center = ESC + b'a' + b'\x01'  # Center alignment
        raster_cmd = GS + b'v0' + b'\x00'  # GS v 0 mode 0
        nL = width // 8 % 256
        nH = width // 8 // 256
        mL = height % 256
        mH = height // 256
        header = raster_cmd + bytes([nL, nH, mL, mH])
        
        # Convert image to bytes
        bitmap = bytearray()
        for y in range(height):
            for x in range(0, width, 8):
                byte = 0
                for bit in range(8):
                    if x + bit < width and pixels[x + bit, y] == 0:
                        byte |= 1 << (7 - bit)
                bitmap.append(byte)
        
        # Log bitmap size
        logging.info(f"Generated bitmap size: {len(bitmap)} bytes")
        
        # Open printer
        printer_name = win32print.GetDefaultPrinter()
        hPrinter = win32print.OpenPrinter(printer_name)
        try:
            hJob = win32print.StartDocPrinter(hPrinter, 1, ("Invoice", None, "RAW"))
            win32print.StartPagePrinter(hPrinter)
            
            # Send ESC/POS commands and image data
            win32print.WritePrinter(hPrinter, init_cmd)
            win32print.WritePrinter(hPrinter, align_center)
            win32print.WritePrinter(hPrinter, header)
            win32print.WritePrinter(hPrinter, bitmap)
            
            # Paper feed and cut
            win32print.WritePrinter(hPrinter, b'\n\n\n')
            win32print.WritePrinter(hPrinter, GS + b'V' + b'\x00')  # GS V 0 (full cut)
            
            win32print.EndPagePrinter(hPrinter)
            win32print.EndDocPrinter(hPrinter)
            logging.info("Successfully sent data to printer")
        except Exception as e:
            logging.error(f"Error sending image to printer: {e}")
            raise
        finally:
            win32print.ClosePrinter(hPrinter)

class InvoiceWindow(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("G.V. Mahant Brothers - Invoice")
        self.geometry("1200x800")
        self.configure(fg_color=BACKGROUND_COLOR)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.main_frame = ctk.CTkFrame(self, fg_color=BACKGROUND_COLOR)
        self.main_frame.grid(row=0, column=0, sticky="nsew")
        self.main_frame.grid_rowconfigure(4, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        self._setup_header()
        self._setup_mode_selection()
        self._setup_customer_section()
        self._setup_table()
        self._setup_bottom_section()
        self.switch_mode("Patti")

    def _setup_header(self):
        self.header_frame = ctk.CTkFrame(self.main_frame, fg_color=PRIMARY_COLOR)
        self.header_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 0))
        self.header_frame.grid_columnconfigure(0, weight=1)
        shree_label = ctk.CTkLabel(self.header_frame, text="|ಶ್ರೀ|", font=("Noto Sans Kannada", 28, "bold"), text_color=BACKGROUND_COLOR, anchor="w", justify="left")
        shree_label.grid(row=0, column=0, sticky="w", padx=20, pady=(10, 0))
        company_label = ctk.CTkLabel(self.header_frame, text="G.V. Mahant Brothers", font=HEADER_FONT, text_color=BACKGROUND_COLOR, anchor="w", justify="left")
        company_label.grid(row=1, column=0, sticky="w", padx=20, pady=(0, 10))
        self.date_label = ctk.CTkLabel(self.header_frame, font=SUBHEADER_FONT, text_color=BACKGROUND_COLOR, anchor="e", justify="right")
        self.date_label.grid(row=0, column=1, rowspan=2, sticky="e", padx=20, pady=10)
        self._update_datetime()

    def _update_datetime(self):
        now = datetime.now()
        self.date_label.configure(text=now.strftime("%A, %d %B %Y\n%I:%M %p"))
        self.after(1000, self._update_datetime)

    def _setup_mode_selection(self):
        self.mode_frame = ctk.CTkFrame(self.main_frame, fg_color=FRAME_COLOR)
        self.mode_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(10, 0))
        self.mode_var = ctk.StringVar(value="Patti")
        btn_container = ctk.CTkFrame(self.mode_frame, fg_color=FRAME_COLOR)
        btn_container.pack(expand=True)
        for mode in ["Patti", "Kata", "Barthe"]:
            btn = ctk.CTkRadioButton(
                btn_container, text=mode, variable=self.mode_var, value=mode,
                font=LABEL_FONT, fg_color=PRIMARY_COLOR, hover_color=SECONDARY_COLOR,
                border_color=BORDER_COLOR, text_color=TEXT_COLOR, command=self._on_mode_change
            )
            btn.pack(side="left", padx=8, pady=8)

    def _on_mode_change(self):
        self.switch_mode(self.mode_var.get())

    def switch_mode(self, mode):
        self.current_mode = mode
        self.table.set_mode(mode)
        self._setup_kata_field()
        self.table.update_amounts()

    def _setup_customer_section(self):
        self.customer_frame = ctk.CTkFrame(self.main_frame, fg_color=BACKGROUND_COLOR)
        self.customer_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=(10, 0))
        ctk.CTkLabel(self.customer_frame, text="Customer Name:", font=LABEL_FONT, text_color=TEXT_COLOR).pack(side="left", padx=10)
        self.customer_entry = ctk.CTkEntry(self.customer_frame, font=ENTRY_FONT, width=400, placeholder_text="Enter customer name")
        self.customer_entry.pack(side="left", padx=10)
        ctk.CTkLabel(self.customer_frame, text="", fg_color=BACKGROUND_COLOR).pack(side="left", expand=True, fill="x")

    def _setup_table(self):
        self.table = InvoiceTable(self.main_frame, mode="Patti")
        self.table.grid(row=4, column=0, sticky="nsew", padx=10, pady=(10, 0))

    def _setup_bottom_section(self):
        self.bottom_frame = ctk.CTkFrame(self.main_frame, fg_color=BACKGROUND_COLOR)
        self.bottom_frame.grid(row=5, column=0, sticky="ew", padx=10, pady=10)
        add_row_btn = ctk.CTkButton(
            self.bottom_frame,
            text="Add Row",
            command=self.table.add_row,
            fg_color=PRIMARY_COLOR,
            hover_color=SECONDARY_COLOR,
            font=BUTTON_FONT
        )
        add_row_btn.pack(side="left", padx=5)
        clear_btn = ctk.CTkButton(
            self.bottom_frame,
            text="Clear",
            command=self.table.clear_rows,
            fg_color=PRIMARY_COLOR,
            hover_color=SECONDARY_COLOR,
            font=BUTTON_FONT
        )
        clear_btn.pack(side="left", padx=5)
        save_btn = ctk.CTkButton(
            self.bottom_frame,
            text="Save",
            command=self.save_to_excel,
            fg_color=PRIMARY_COLOR,
            hover_color=SECONDARY_COLOR,
            font=BUTTON_FONT
        )
        save_btn.pack(side="left", padx=5)
        
        print_btn = ctk.CTkButton(
            self.bottom_frame,
            text="Print",
            command=self.show_print_preview,
            fg_color=PRIMARY_COLOR,
            hover_color=SECONDARY_COLOR,
            font=BUTTON_FONT
        )
        print_btn.pack(side="left", padx=5)
        
        self.right_bottom_frame = ctk.CTkFrame(self.bottom_frame, fg_color=BACKGROUND_COLOR)
        self.right_bottom_frame.pack(side="right", padx=5, pady=0, anchor="se")
        self.kata_frame = ctk.CTkFrame(self.right_bottom_frame, fg_color=BACKGROUND_COLOR)
        self.kata_frame.pack(side="top", pady=(0, 2), anchor="e")
        self.total_label = ctk.CTkLabel(
            self.right_bottom_frame,
            text="Amount: ₹0.00",
            text_color=PRIMARY_COLOR,
            font=HEADER_FONT
        )
        self.total_label.pack(side="top", anchor="e")

    def _setup_kata_field(self):
        for w in self.kata_frame.winfo_children():
            w.destroy()
        if self.current_mode == "Kata":
            ctk.CTkLabel(self.kata_frame, text="", fg_color=BACKGROUND_COLOR).pack(side="top", pady=(10, 0))
            kata_label = ctk.CTkLabel(self.kata_frame, text="Kata:", font=LABEL_FONT_BOLD, text_color=TEXT_COLOR)
            kata_label.pack(side="left", padx=(0, 18), pady=(12, 12))
            self.kata_amount_entry = ctk.CTkEntry(self.kata_frame, font=ENTRY_FONT, width=120)
            self.kata_amount_entry.insert(0, "0")
            self.kata_amount_entry.pack(side="left", padx=(0, 10), pady=(12, 12))
            self.kata_amount_entry.bind("<KeyRelease>", lambda e: self.table.update_amounts())
        else:
            self.kata_amount_entry = None

    def save_to_excel(self):
        try:
            home_dir = os.path.expanduser("~")
            documents_path = os.path.join(home_dir, "Documents")
            os.makedirs(documents_path, exist_ok=True)
            date_str = datetime.now().strftime('%Y-%m-%d')
            base_filename = f"Invoice_{date_str}.xlsx"
            full_save_path = os.path.join(documents_path, base_filename)
            customer = self.customer_entry.get().strip() or "Unknown Customer"
            mode = self.current_mode
            headers = MODE_HEADERS[mode]
            data_rows = self.table.get_all_data()
            if not data_rows:
                messagebox.showwarning("No Data", "No data entered to save.")
                return
            try:
                if os.path.exists(full_save_path):
                    wb = load_workbook(full_save_path)
                else:
                    wb = Workbook()
                if mode in wb.sheetnames:
                    ws = wb[mode]
                else:
                    if len(wb.sheetnames) > 0:
                        ws = wb.create_sheet(title=mode)
                    else:
                        ws = wb.active
                        ws.title = mode
                ws.delete_rows(1, ws.max_row)
                ws.append(["Timestamp", "Customer"] + headers)
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                for row in data_rows:
                    ws.append([timestamp, customer] + row)
                wb.save(full_save_path)
                messagebox.showinfo("Saved", f"Invoice data saved to:\n{full_save_path}\n(Sheet: {mode})")
            except PermissionError:
                messagebox.showerror("Permission Error", f"Cannot save '{base_filename}'.\nThe file might be open in Excel.\n\nLocation: {documents_path}")
            except Exception as e:
                messagebox.showerror("Save Error", f"Error saving Excel file to:\n{full_save_path}\n\nError: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Unexpected error during save operation: {str(e)}")

    def show_print_preview(self):
        """Show a preview of the invoice before printing."""
        try:
            customer_name = self.customer_entry.get().strip() or ""
            data_rows = self.table.get_all_data()
            
            if not data_rows:
                messagebox.showwarning("No Data", "No data to print.")
                return

            # Prepare content with 48-character width
            max_width = 48
            content = []
            dynamic_content = []
            
            # Header (in template)
            content.append("=" * max_width)
            content.append("ಶ್ರೀ".center(max_width))
            content.append("G.V. Mahant Brothers".center(max_width))
            content.append("Invoice".center(max_width))
            content.append("=" * max_width)
            
            # Date, Time, Customer
            now = datetime.now()
            date_str = now.strftime("%d/%m/%Y")
            time_str = now.strftime("%I:%M %p")
            content.append(f"Date: {date_str}")
            content.append(f"Time: {time_str}")
            content.append(f"Customer: {customer_name[:30]}")
            content.append("-" * max_width)
            dynamic_content.extend([
                f"Date: {date_str}",
                f"Time: {time_str}",
                f"Customer: {customer_name[:30]}",
                "-" * max_width
            ])
            
            # Table headers
            headers = MODE_HEADERS[self.current_mode]
            col_widths = [12, 8, 8, 8, 8, 10] if self.current_mode in ["Patti", "Kata"] else [12, 8, 8, 8, 8, 8, 10]
            header_line = "".join(f"{h:<{w}}" for h, w in zip(headers[:-1], col_widths))
            content.append(header_line.rstrip())
            content.append("-" * max_width)
            dynamic_content.extend([
                header_line.rstrip(),
                "-" * max_width
            ])
            
            # Table data
            total_amount = 0
            for row in data_rows:
                if any(str(cell).strip() for cell in row):
                    amount = float(row[-2].replace("₹", "")) if row[-2].startswith("₹") else 0
                    row_str = "".join(f"{str(cell):<{w}}" for cell, w in zip(row[:-1], col_widths[:len(row)-1]))
                    content.append(row_str.rstrip())
                    dynamic_content.append(row_str.rstrip())
                    total_amount += amount
            
            content.append("-" * max_width)
            dynamic_content.append("-" * max_width)
            
            # Kata amount (if in Kata mode)
            if self.current_mode == "Kata" and hasattr(self, 'kata_amount_entry') and self.kata_amount_entry:
                kata_amount = safe_float(self.kata_amount_entry.get())
                if kata_amount > 0:
                    content.append(f"Kata Amount: ₹{kata_amount:.2f}".rstrip())
                    content.append("-" * max_width)
                    dynamic_content.extend([
                        f"Kata Amount: ₹{kata_amount:.2f}".rstrip(),
                        "-" * max_width
                    ])
                    total_amount += kata_amount
            
            # Total amount
            content.append(f"Amount: ₹{total_amount:.2f}")
            content.append("=" * max_width)
            dynamic_content.extend([
                f"Amount: ₹{total_amount:.2f}",
                "=" * max_width
            ])
            
            # Footer (in template)
            footer = "ನಾನು ಎಲ್ಲವೂ ಸರಿಯಾಗಿದೆ ಎಂದು ಪರಿಶೀಲಿಸಿದ್ದೇನೆ."
            content.append(footer.center(max_width))
            content.append("")  # Extra line for paper feed
            
            # Join content for preview
            preview_content = '\n'.join(content)
            
            # Show the preview dialog
            PrintPreviewDialog(self, preview_content, self.current_mode, dynamic_content)
            
        except Exception as e:
            logging.error(f"Error showing print preview: {e}")
            messagebox.showerror("Preview Error", f"Error showing print preview: {str(e)}")

def test_font_support(font, test_text):
    """Test if the font can render the given text without boxes."""
    try:
        # Create a small image to test rendering
        img = Image.new('1', (100, 100), 1)  # White background
        draw = ImageDraw.Draw(img)
        draw.text((0, 0), test_text, font=font, fill=0)  # Black text
        
        # Convert image to numpy array to check for rendered content
        img_array = np.array(img)
        # If the image is all white (value 1), the text didn't render (missing glyphs)
        if np.all(img_array == 1):
            logging.warning(f"Font failed to render test text '{test_text}' (all white).")
            return False
        # If there are black pixels (value 0), the text rendered successfully
        if np.any(img_array == 0):
            return True
        else:
            logging.warning(f"Font failed to render test text '{test_text}' (no black pixels).")
            return False
    except Exception as e:
        logging.warning(f"Font failed to render test text '{test_text}': {e}")
        return False

if __name__ == "__main__":
    app = InvoiceWindow()
    app.mainloop()