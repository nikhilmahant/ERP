import sys
import os
import json
import logging
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QFrame, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox, QRadioButton,
    QButtonGroup, QSpacerItem, QSizePolicy, QDialog, QTextEdit, QFileDialog
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QPalette, QColor, QFont
from constants import *
from openpyxl import Workbook, load_workbook
import win32print
import win32api

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def validate_float(value):
    """Validate if a string can be converted to float."""
    try:
        return float(value) if value.strip() else 0
    except ValueError:
        return 0

class InvoiceWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("G.V. Mahant Brothers - Invoice")
        self.setMinimumSize(1200, 800)
        # Set window background
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f8f9fa;
            }
        """)
        self.load_config()
        self.setup_ui()

    def load_config(self):
        """Load configuration from config.json file."""
        try:
            config_path = os.path.join(os.path.dirname(__file__), 'config.json')
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    self.config = json.load(f)
            else:
                # Default configuration
                self.config = {
                    'items': [],
                    'last_customer': '',
                    'last_mode': 'Patti'
                }
                # Save default config
                with open(config_path, 'w') as f:
                    json.dump(self.config, f, indent=4)
        except Exception as e:
            logging.error(f"Error loading config: {e}")
            self.config = {
                'items': [],
                'last_customer': '',
                'last_mode': 'Patti'
            }

    def setup_ui(self):
        """Set up the main UI components."""
        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # Create top section for customer and mode selection
        top_section = QWidget()
        top_layout = QHBoxLayout(top_section)
        
        # Customer entry
        customer_label = QLabel("Customer:")
        customer_label.setFont(LABEL_FONT)
        self.customer_entry = QLineEdit()
        self.customer_entry.setFont(ENTRY_FONT)
        self.customer_entry.setText(self.config.get('last_customer', ''))
        top_layout.addWidget(customer_label)
        top_layout.addWidget(self.customer_entry)
        
        # Mode selection
        mode_label = QLabel("Mode:")
        mode_label.setFont(LABEL_FONT)
        self.mode_group = QButtonGroup()
        self.patti_radio = QRadioButton("Patti")
        self.kata_radio = QRadioButton("Kata")
        self.barthe_radio = QRadioButton("Barthe")
        
        for radio in [self.patti_radio, self.kata_radio, self.barthe_radio]:
            radio.setFont(ENTRY_FONT)
            self.mode_group.addButton(radio)
            top_layout.addWidget(radio)
        
        # Set default mode
        last_mode = self.config.get('last_mode', 'Patti')
        if last_mode == "Patti":
            self.patti_radio.setChecked(True)
        elif last_mode == "Kata":
            self.kata_radio.setChecked(True)
        elif last_mode == "Barthe":
            self.barthe_radio.setChecked(True)
        
        self.current_mode = last_mode
        self.mode_group.buttonClicked.connect(self.on_mode_changed)
        
        # Add top section to main layout
        main_layout.addWidget(top_section)
        
        # Create table
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Item", "Pkt/Net", "Qty/Less%", "Rate", "Hamali", "Amount"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        main_layout.addWidget(self.table)
        
        # Create bottom section for total and buttons
        bottom_section = QWidget()
        bottom_layout = QHBoxLayout(bottom_section)
        
        # Total amount
        self.total_label = QLabel("Total: ₹0.00")
        self.total_label.setFont(LABEL_FONT)
        bottom_layout.addWidget(self.total_label)
        
        # Buttons
        add_row_btn = QPushButton("Add Row")
        add_row_btn.clicked.connect(self.add_row)
        print_btn = QPushButton("Print")
        print_btn.clicked.connect(self.save_for_print)
        save_btn = QPushButton("Save")
        save_btn.clicked.connect(self.save_to_excel)
        
        for btn in [add_row_btn, print_btn, save_btn]:
            btn.setFont(BUTTON_FONT)
            bottom_layout.addWidget(btn)
        
        main_layout.addWidget(bottom_section)
        
        # Create kata frame (initially hidden)
        self.kata_frame = QFrame()
        self.kata_layout = QHBoxLayout(self.kata_frame)
        self.kata_frame.hide()
        
        # Add initial row
        self.add_row()

    def on_mode_changed(self, button):
        """Handle mode change and update UI accordingly."""
        self.current_mode = button.text()
        self.config['last_mode'] = self.current_mode
        
        # Update table headers based on mode
        if self.current_mode == "Patti":
            self.table.setHorizontalHeaderLabels(["Item", "Pkt", "Qty", "Rate", "Hamali", "Amount"])
        elif self.current_mode == "Kata":
            self.table.setHorizontalHeaderLabels(["Item", "Net", "Less%", "Rate", "Hamali", "Amount"])
        elif self.current_mode == "Barthe":
            self.table.setHorizontalHeaderLabels(["Item", "Pkt", "Wt", "Adj", "Rate", "Hamali", "Amount"])
        
        # Update kata field visibility
        self.setup_kata_field()
        
        # Update amounts
        self.update_amounts()

    def add_row(self):
        """Add a new row to the table with appropriate widgets."""
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # Item combo box
        item_combo = QComboBox()
        item_combo.setFont(ENTRY_FONT)
        item_combo.addItems(self.config.get('items', []))
        item_combo.setEditable(True)
        self.table.setCellWidget(row, 0, item_combo)
        
        # Create line edits for other columns
        for col in range(1, 6):
            line_edit = QLineEdit()
            line_edit.setFont(ENTRY_FONT)
            line_edit.setText("0")
            line_edit.textChanged.connect(self.update_amounts)
            self.table.setCellWidget(row, col, line_edit)
        
        # Connect item combo box change
        item_combo.currentTextChanged.connect(self.update_amounts)

    def update_amounts(self):
        """Update the amount column and total for each row."""
        total_amount = 0
        
        for row in range(self.table.rowCount()):
            try:
                if self.current_mode == "Patti":
                    packet = float(self.table.cellWidget(row, 1).text() or 0)
                    quantity = float(self.table.cellWidget(row, 2).text() or 0)
                    rate = float(self.table.cellWidget(row, 3).text() or 0)
                    hamali = float(self.table.cellWidget(row, 4).text() or 0)
                    amount = (quantity * rate) + (packet * hamali)
                elif self.current_mode == "Kata":
                    net_wt = float(self.table.cellWidget(row, 1).text() or 0)
                    less_percent = float(self.table.cellWidget(row, 2).text() or 0)
                    rate = float(self.table.cellWidget(row, 3).text() or 0)
                    hamali_rate = float(self.table.cellWidget(row, 4).text() or 0)
                    final_wt = net_wt * (1 - less_percent / 100.0) if less_percent < 100 else 0.0
                    packets = int(net_wt / 60) if net_wt > 0 else 0
                    amount = (final_wt * rate) + (packets * hamali_rate)
                elif self.current_mode == "Barthe":
                    packet = float(self.table.cellWidget(row, 1).text() or 0)
                    weight = float(self.table.cellWidget(row, 2).text() or 0)
                    adjustment = float(self.table.cellWidget(row, 3).text() or 0)
                    rate = float(self.table.cellWidget(row, 4).text() or 0)
                    hamali = float(self.table.cellWidget(row, 5).text() or 0)
                    total_qty = (packet * weight) + adjustment
                    amount = (total_qty * rate) + (packet * hamali)
                
                # Update amount cell
                amount_widget = self.table.cellWidget(row, 5)
                amount_widget.setText(f"{amount:.2f}")
                total_amount += amount
                
            except Exception as e:
                logging.error(f"Error updating row {row}: {str(e)}")
        
        # Add kata amount if in Kata mode
        if self.current_mode == "Kata" and hasattr(self, 'kata_amount_entry'):
            try:
                kata_amount = float(self.kata_amount_entry.text() or 0)
                total_amount += kata_amount
            except Exception as e:
                logging.error(f"Error processing kata amount: {str(e)}")
        
        # Update total label
        self.total_label.setText(f"Total: ₹{total_amount:.2f}")

    def save_to_excel(self):
        """Save the current invoice data to an Excel file."""
        try:
            # Get save file path
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Invoice",
                os.path.join(os.path.dirname(__file__), f"invoice_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
                
            # Create workbook and select active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice"
            
            # Write header
            ws['A1'] = "G.V. Mahant Brothers"
            ws['A2'] = datetime.now().strftime("%d-%b-%Y %I:%M %p")
            ws['A3'] = f"Customer: {self.customer_entry.text()}"
            ws['A4'] = f"Mode: {self.current_mode}"
            
            # Write table headers
            headers = ["Item", "Pkt/Net", "Qty/Less%", "Rate", "Hamali", "Amount"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=6, column=col, value=header)
            
            # Write data rows
            for row in range(self.table.rowCount()):
                for col in range(6):
                    widget = self.table.cellWidget(row, col)
                    if widget:
                        value = widget.text() if isinstance(widget, QLineEdit) else widget.currentText()
                        ws.cell(row=row+7, column=col+1, value=value)
            
            # Add kata amount if in Kata mode
            if self.current_mode == "Kata" and hasattr(self, 'kata_amount_entry'):
                kata_row = self.table.rowCount() + 7
                ws.cell(row=kata_row, column=1, value="Kata Amount")
                ws.cell(row=kata_row, column=6, value=self.kata_amount_entry.text())
            
            # Add total
            total_row = self.table.rowCount() + 8
            ws.cell(row=total_row, column=1, value="Total")
            ws.cell(row=total_row, column=6, value=self.total_label.text().replace("Total: ₹", ""))
            
            # Save workbook
            wb.save(file_path)
            QMessageBox.information(self, "Success", "Invoice saved successfully!")
            
        except Exception as e:
            logging.error(f"Error saving to Excel: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to save invoice: {str(e)}")

    def generate_print_content(self):
        """Generates the formatted string list for printing/preview, matching the Tkinter format."""
        logging.info("Generating print content...")
        lines = []
        
        # Validate customer name
        customer = self.customer_entry.text().strip()
        if not customer:
            QMessageBox.warning(self, "Warning", "Customer name is empty. Using 'Unknown Customer'.")
            customer = "Unknown Customer"
            
        mode = self.current_mode
        logging.info(f"Generating content for mode: {mode}, customer: {customer}")

        # --- Header ---
        lines.append("          G.V. Mahant Brothers          ")
        lines.append(datetime.now().strftime("          %d-%b-%Y %I:%M %p          "))
        lines.append("-" * 48)
        lines.append(f"Customer: {customer}")
        lines.append("-" * 48)

        # --- Column Headers ---
        if mode == "Patti":
            lines.append(f"{'Item':<10}    {'Pkt':>3}  {'Qty':>3}  {'Rate':>5}   {'Ham':>3}      {'Amt':>5}")
        elif mode == "Kata":
            lines.append(f"{'Item':<10}    {'Net':>3}  {'Less%':>5}  {'Rate':>5}   {'Ham':>3}      {'Amt':>5}")
        elif mode == "Barthe":
            lines.append(f"{'Item':<10}    {'Pkt':>3}  {'Wt':>3}  {'Adj':>3}  {'Rate':>5}   {'Ham':>3}      {'Amt':>5}")

        lines.append("-" * 48)

        # --- Data Rows ---
        total_amount = 0
        row_count = 0
        for row in range(self.table.rowCount()):
            try:
                item_combo = self.table.cellWidget(row, 0)
                if not item_combo or not item_combo.currentText().strip():
                    continue
                    
                row_count += 1
                if mode == "Patti":
                    item = item_combo.currentText()[:10]
                    packet = float(self.table.cellWidget(row, 1).text() or 0)
                    quantity = float(self.table.cellWidget(row, 2).text() or 0)
                    rate = float(self.table.cellWidget(row, 3).text() or 0)
                    hamali = float(self.table.cellWidget(row, 4).text() or 0)
                    amount = (quantity * rate) + (packet * hamali)
                    lines.append(f"{item:<10}    {packet:>3.1f}  {quantity:>3.1f}  {rate:>5.1f}   {hamali:>3.1f}    {amount:>7.2f}")
                elif mode == "Kata":
                    item = item_combo.currentText()[:10]
                    net_wt = float(self.table.cellWidget(row, 1).text() or 0)
                    less_percent = float(self.table.cellWidget(row, 2).text() or 0)
                    rate = float(self.table.cellWidget(row, 3).text() or 0)
                    hamali_rate = float(self.table.cellWidget(row, 4).text() or 0)
                    final_wt = net_wt * (1 - less_percent / 100.0) if less_percent < 100 else 0.0
                    packets = int(net_wt / 60) if net_wt > 0 else 0
                    amount = (final_wt * rate) + (packets * hamali_rate)
                    lines.append(f"{item:<10}    {net_wt:>3.1f}  {less_percent:>5.1f}  {rate:>5.1f}   {hamali_rate:>3.1f}    {amount:>7.2f}")
                elif mode == "Barthe":
                    item = item_combo.currentText()[:10]
                    packet = float(self.table.cellWidget(row, 1).text() or 0)
                    weight = float(self.table.cellWidget(row, 2).text() or 0)
                    adjustment = float(self.table.cellWidget(row, 3).text() or 0)
                    rate = float(self.table.cellWidget(row, 4).text() or 0)
                    hamali = float(self.table.cellWidget(row, 5).text() or 0)
                    total_qty = (packet * weight) + adjustment
                    amount = (total_qty * rate) + (packet * hamali)
                    lines.append(f"{item:<10}    {packet:>3.1f}  {weight:>3.1f}  {adjustment:>3.1f}  {rate:>5.1f}   {hamali:>3.1f}    {amount:>7.2f}")
                total_amount += amount
            except Exception as e:
                logging.error(f"Error processing row {row}: {str(e)}")
                lines.append(f"Error: {str(e)}")

        if row_count == 0:
            logging.warning("No valid rows found in the table")
            lines.append("No items to print")
            lines.append("-" * 48)

        # Add kata amount if in Kata mode
        if mode == "Kata" and hasattr(self, 'kata_amount_entry'):
            try:
                kata_amount = float(self.kata_amount_entry.text() or 0)
                total_amount += kata_amount
                lines.append("-" * 48)
                lines.append(f"{'':>14}Kata Amount: {kata_amount:>7.2f}")
            except Exception as e:
                logging.error(f"Error processing kata amount: {str(e)}")
                lines.append(f"Error: {str(e)}")

        lines.append("-" * 48)
        lines.append(f"{'':>14}Total Amount: {total_amount:>7.2f}")
        lines.append("-" * 48)
        lines.append("ನಾನು ಎಲ್ಲವೂ ಸರಿಯಾಗಿದೆ ಎಂದು ಪರಿಶೀಲಿಸಿದ್ದೇನೆ.")
        lines.append("")
        lines.append("")
        lines.append("")
        
        logging.info(f"Generated {len(lines)} lines for printing")
        return lines

    def save_for_print(self):
        """Prints the generated content to the default printer using robust encoding and cut command."""
        # Validate table data
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No data to print.")
            return

        try:
            printer_name = win32print.GetDefaultPrinter()
            if not printer_name:
                QMessageBox.critical(self, "Printer Error", "No default printer found!")
                return
                
            logging.info(f"Attempting to print to: {printer_name}")
            hPrinter = win32print.OpenPrinter(printer_name)
            
            try:
                # ESC/POS: Initialize printer
                init_printer = b'\x1B\x40'
                # Start print job in RAW mode
                hJob = win32print.StartDocPrinter(hPrinter, 1, ("Invoice", None, "RAW"))
                win32print.StartPagePrinter(hPrinter)
                
                # Send initialization command
                win32print.WritePrinter(hPrinter, init_printer)
                
                # Print content line by line with proper encoding
                lines = self.generate_print_content()
                logging.info(f"Printing {len(lines)} lines...")
                
                for i, line in enumerate(lines, 1):
                    try:
                        encoded_line = line.encode('ascii') + b'\n'
                    except UnicodeEncodeError:
                        encoded_line = line.encode('utf-8') + b'\n'
                    win32print.WritePrinter(hPrinter, encoded_line)
                    win32api.Sleep(50)  # Small delay between lines
                    if i % 10 == 0:  # Log progress every 10 lines
                        logging.info(f"Printed {i} of {len(lines)} lines")
                
                # Paper feed and cut commands
                win32print.WritePrinter(hPrinter, b'\n\n\n\n')  # Feed 4 lines
                win32print.WritePrinter(hPrinter, b'\x1D\x56\x00')  # ESC/POS cut command
                win32api.Sleep(300)  # Wait for cut to complete
                
                win32print.EndPagePrinter(hPrinter)
                win32print.EndDocPrinter(hPrinter)
                logging.info("Print job completed successfully")
                QMessageBox.information(self, "Success", "Printed successfully with paper cut!")
                
            except Exception as e:
                logging.error(f"Printing error: {e}")
                QMessageBox.critical(self, "Print Error", f"Failed to print: {str(e)}")
            finally:
                win32print.ClosePrinter(hPrinter)
                logging.info("Printer handle closed")
                
        except Exception as e:
            logging.error(f"Printer connection error: {e}")
            QMessageBox.critical(self, "Printer Error", f"Cannot connect to printer: {str(e)}")

    def setup_kata_field(self):
        # Clear existing kata field
        for i in reversed(range(self.kata_layout.count())): 
            self.kata_layout.itemAt(i).widget().setParent(None)
        
        # Remove kata_frame from layout if it exists
        if self.kata_frame.parent():
            self.total_layout.removeWidget(self.kata_frame)
            self.kata_frame.hide()
        
        # Remove all widgets from total_layout
        while self.total_layout.count():
            item = self.total_layout.takeAt(0)
            if item.widget():
                item.widget().hide()
        
        # Only show kata amount field in Kata mode
        if self.current_mode == "Kata":
            kata_label = QLabel("Kata:")
            kata_label.setFont(LABEL_FONT)
            self.kata_amount_entry = QLineEdit()
            self.kata_amount_entry.setFont(ENTRY_FONT)
            self.kata_amount_entry.setFixedWidth(120)
            self.kata_amount_entry.setText("0")
            self.kata_amount_entry.textChanged.connect(self.update_amounts)
            
            self.kata_layout.addWidget(kata_label)
            self.kata_layout.addWidget(self.kata_amount_entry)
            
            # Insert kata_frame at the beginning of total_layout
            self.kata_frame.show()
            self.total_layout.insertWidget(0, self.kata_frame)
        
        # Always show total label
        self.total_label.show()
        self.total_layout.addWidget(self.total_label)
        self.update_amounts() 