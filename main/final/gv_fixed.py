import sys
import os
import json
import logging
from datetime import datetime
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QComboBox, QFrame, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox, QRadioButton,
    QButtonGroup, QSpacerItem, QSizePolicy, QDialog, QTextEdit
)
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QPalette, QColor
from constants import *
from openpyxl import Workbook, load_workbook
import win32print
import win32api

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def validate_float(value):
    """Validate if a string can be converted to float."""
    try:
        return float(value) if value.strip() else 0
    except ValueError:
        return 0

class InvoiceWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("G.V. Mahant Brothers - Invoice")
        self.setMinimumSize(1200, 800)
        self.load_config()
        self.setup_ui()

    def clear_rows(self):
        self.table.setRowCount(0)
        self.add_row()
        self.update_amounts()

    def load_config(self):
        """Load application configuration from file"""
        self.config = {
            "theme": "Light",
            "window_size": "1200x800",
            "autosave": True
        }
        try:
            if os.path.exists(CONFIG_FILE):
                with open(CONFIG_FILE, 'r') as f:
                    self.config.update(json.load(f))
        except Exception as e:
            logging.error(f"Error loading config: {e}")

    def manage_items(self):
        QMessageBox.information(self, "Manage Items", "This feature is not implemented yet.")

    def setup_ui(self):
        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)  # Spacing between major sections

        # Header section
        self.setup_header()
        
        # Mode selection
        self.setup_mode_selection()
        
        # Customer section
        self.setup_customer_section()
        
        # Table section
        self.setup_table()
        
        # Bottom section
        self.setup_bottom_section()

        # Initialize the first row
        self.add_row()

    def setup_header(self):
        header_frame = QFrame()
        header_frame.setStyleSheet("QFrame { background-color: #4682b4; border-radius: 10px; }")  # Steel blue background
        header_layout = QVBoxLayout(header_frame)
        header_layout.setContentsMargins(15, 15, 15, 15)
        header_layout.setSpacing(10)
        
        # Title labels
        shree_label = QLabel("ಶ್ರೀ")
        shree_label.setFont(QFont("Segoe UI", 24, QFont.Bold))
        shree_label.setAlignment(Qt.AlignCenter)
        shree_label.setStyleSheet("color: #ffffff;")  # White color for better visibility
        
        title_label = QLabel("G.V. Mahant Brothers")
        title_label.setFont(QFont("Segoe UI", 20, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("color: #ffffff;")  # White color for better visibility
        
        date_label = QLabel(datetime.now().strftime("%A, %d %B %Y %I:%M %p"))
        date_label.setFont(QFont("Segoe UI", 12))
        date_label.setAlignment(Qt.AlignCenter)
        date_label.setStyleSheet("color: #ffffff;")  # White color for better visibility
        
        header_layout.addWidget(shree_label)
        header_layout.addWidget(title_label)
        header_layout.addWidget(date_label)
        
        self.centralWidget().layout().addWidget(header_frame)

    def setup_mode_selection(self):
        mode_frame = QFrame()
        mode_frame.setStyleSheet("""
            QFrame { 
                background-color: #e9ecef;
                border-radius: 8px;
                padding: 10px;
            }
            QRadioButton {
                padding: 8px 15px;
                border-radius: 5px;
                min-width: 100px;
                text-align: center;
            }
            QRadioButton:checked {
                background-color: #0d6efd;
                color: white;
            }
        """)
        mode_layout = QHBoxLayout(mode_frame)
        mode_layout.setContentsMargins(20, 10, 20, 10)
        mode_layout.setSpacing(20)
        mode_layout.addStretch()
        
        self.mode_group = QButtonGroup(self)
        self.current_mode = "Patti"  # Default mode
        
        for mode in ["Patti", "Kata", "Barthe"]:
            radio = QRadioButton(mode)
            radio.setFont(QFont("Segoe UI", 12))
            if mode == "Patti":
                radio.setChecked(True)
            radio.toggled.connect(self.switch_mode)
            self.mode_group.addButton(radio)
            mode_layout.addWidget(radio)
        
        mode_layout.addStretch()
        self.centralWidget().layout().addWidget(mode_frame)

    def setup_customer_section(self):
        customer_frame = QFrame()
        customer_frame.setStyleSheet("""
            QFrame { 
                background-color: #f8f9fa;
                border-radius: 8px;
                padding: 10px;
            }
            QLineEdit {
                padding: 8px;
                border: 2px solid #ced4da;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus {
                border-color: #86b7fe;
                outline: 0;
                box-shadow: 0 0 0 0.25rem rgba(13, 110, 253, 0.25);
            }
        """)
        customer_layout = QHBoxLayout(customer_frame)
        customer_layout.setContentsMargins(20, 15, 20, 15)
        
        customer_label = QLabel("Customer Name:")
        customer_label.setFont(QFont("Segoe UI", 12))
        customer_label.setStyleSheet("color: #212529;")
        
        self.customer_entry = QLineEdit()
        self.customer_entry.setFont(QFont("Segoe UI", 12))
        self.customer_entry.setMinimumWidth(400)
        self.customer_entry.setPlaceholderText("Enter customer name")
        
        customer_layout.addWidget(customer_label)
        customer_layout.addWidget(self.customer_entry)
        customer_layout.addStretch()
        
        self.centralWidget().layout().addWidget(customer_frame)

    def setup_table(self):
        table_frame = QFrame()
        table_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 8px;
            }
            QTableWidget {
                border: none;
                gridline-color: #dee2e6;
                selection-background-color: #e9ecef;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 8px;
                border: none;
                border-bottom: 2px solid #dee2e6;
                font-weight: bold;
                color: #495057;
            }
        """)
        table_layout = QVBoxLayout(table_frame)
        table_layout.setContentsMargins(1, 1, 1, 1)
        
        self.table = QTableWidget()
        self.table.setFont(QFont("Segoe UI", 11))
        self.update_table_headers()
        
        # Set table properties
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        
        table_layout.addWidget(self.table)
        self.centralWidget().layout().addWidget(table_frame)

    def setup_bottom_section(self):
        bottom_frame = QFrame()
        bottom_layout = QHBoxLayout(bottom_frame)
        bottom_layout.setContentsMargins(0, 10, 0, 10)
        
        # Buttons
        button_frame = QFrame()
        button_frame.setStyleSheet("""
            QPushButton {
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                color: white;
            }
            QPushButton#add { background-color: #198754; }
            QPushButton#add:hover { background-color: #157347; }
            QPushButton#clear { background-color: #dc3545; }
            QPushButton#clear:hover { background-color: #bb2d3b; }
            QPushButton#save { background-color: #0d6efd; }
            QPushButton#save:hover { background-color: #0b5ed7; }
            QPushButton#print { background-color: #6c757d; }
            QPushButton#print:hover { background-color: #5c636a; }
        """)
        button_layout = QHBoxLayout(button_frame)
        button_layout.setSpacing(15)
        
        add_row_btn = QPushButton("Add Row")
        add_row_btn.setObjectName("add")
        clear_btn = QPushButton("Clear")
        clear_btn.setObjectName("clear")
        save_btn = QPushButton("Save")
        save_btn.setObjectName("save")
        print_btn = QPushButton("Print")
        print_btn.setObjectName("print")
        manage_items_btn = QPushButton("Manage Items")
        manage_items_btn.setObjectName("manage")
        manage_items_btn.clicked.connect(self.manage_items)
        button_layout.insertWidget(0, manage_items_btn)
        
        add_row_btn.clicked.connect(self.add_row)
        clear_btn.clicked.connect(self.clear_rows)
        save_btn.clicked.connect(self.save_to_excel)
        print_btn.clicked.connect(self.show_print_preview)
        
        # Total section
        total_frame = QFrame()
        total_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border-radius: 8px;
                padding: 10px;
            }
            QLabel {
                color: #0000ff;
                font-weight: bold;
            }
        """)
        total_layout = QHBoxLayout(total_frame)
        total_layout.setContentsMargins(20, 10, 20, 10)
        
        self.kata_frame = QFrame()
        self.kata_layout = QHBoxLayout(self.kata_frame)
        self.kata_layout.setContentsMargins(0, 0, 20, 0)
        
        self.total_label = QLabel("Amount: ₹0.00")
        self.total_label.setFont(QFont("Segoe UI", 18, QFont.Bold))
        
        total_layout.addWidget(self.kata_frame)
        total_layout.addWidget(self.total_label)
        
        bottom_layout.addWidget(button_frame)
        bottom_layout.addStretch()
        bottom_layout.addWidget(total_frame)
        
        self.centralWidget().layout().addWidget(bottom_frame)

    def update_table_headers(self):
        mode = self.current_mode
        if mode == "Patti":
            headers = ["Item", "Packet", "Quantity", "Rate", "Hamali", "Amount"]
        elif mode == "Kata":
            headers = ["Item", "Net Wt", "Less%", "Rate", "Hamali Rate", "Amount"]
        elif mode == "Barthe":
            headers = ["Item", "Packet", "Weight", "+/-", "Rate", "Hamali", "Amount"]
        
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

    def add_row(self):
        current_row = self.table.rowCount()
        self.table.insertRow(current_row)
        
        # Add item combobox
        item_combo = QComboBox()
        item_combo.setStyleSheet("""
            QComboBox {
                padding: 5px;
                border: 1px solid #ced4da;
                border-radius: 4px;
                background: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: url(down_arrow.png);
                width: 12px;
                height: 12px;
            }
            QComboBox:focus {
                border-color: #86b7fe;
                outline: 0;
            }
        """)
        item_combo.addItems(ITEM_LIST)
        item_combo.currentTextChanged.connect(self.update_amounts)
        item_combo.setFont(QFont("Segoe UI", 11))
        self.table.setCellWidget(current_row, 0, item_combo)
        
        # Add other cells
        for col in range(1, self.table.columnCount() - 1):
            item = QLineEdit()
            item.setStyleSheet("""
                QLineEdit {
                    padding: 5px;
                    border: 1px solid #ced4da;
                    border-radius: 4px;
                    background: white;
                }
                QLineEdit:focus {
                    border-color: #86b7fe;
                    outline: 0;
                }
            """)
            item.setFont(QFont("Segoe UI", 11))
            item.textChanged.connect(self.update_amounts)
            self.table.setCellWidget(current_row, col, item)
        
        # Add amount label
        amount_label = QLabel("₹0.00")
        amount_label.setStyleSheet("""
            QLabel {
                color: #198754;
                padding: 5px;
                background: transparent;
            }
        """)
        amount_label.setFont(QFont("Segoe UI", 11, QFont.Bold))
        amount_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.table.setCellWidget(current_row, self.table.columnCount() - 1, amount_label)

    def switch_mode(self):
        self.current_mode = self.mode_group.checkedButton().text()
        self.update_table_headers()
        self.clear_rows()
        self.setup_kata_field()
        self.update_amounts()

    def show_print_preview(self):
        self.save_to_excel()
        preview_dialog = PrintPreviewDialog(self)
        preview_dialog.exec()

    def generate_print_content(self):
        """Generates the formatted string list for printing/preview."""
        lines = []
        customer = self.customer_entry.text().strip() or "N/A"
        mode = self.current_mode
        max_width = 32  # Adjusted for typical thermal printer width

        # --- Header ---
        lines.extend([
            "|ಶ್ರೀ|".center(max_width),
            "",
            "G.V. Mahant Brothers".center(max_width),
            datetime.now().strftime("%d-%b-%Y %H:%M").center(max_width),
            "-" * max_width,
            f"Customer: {customer}".ljust(max_width),
            "-" * max_width,
        ])

        # --- Column Headers ---
        header_fmt = ""
        if mode == "Patti":
            header_fmt = "{:<5} {:>3} {:>3} {:>5} {:>3} {:>7}"
            headers = ["Item", "Pkt", "Qty", "Rate", "Ham", "Amt"]
        elif mode == "Kata":
            header_fmt = "{:<5} {:>3} {:>3} {:>5} {:>3} {:>7}"
            headers = ["Item", "Net", "Les", "Rate", "Ham", "Amt"]
        elif mode == "Barthe":
            header_fmt = "{:<4} {:>3} {:>3} {:>3} {:>4} {:>3} {:>6}"
            headers = ["Item", "Pkt", "Wt", "+/-", "Rate", "Ham", "Amt"]

        if header_fmt:
            lines.append(header_fmt.format(*headers))
            lines.append("-" * max_width)

        # --- Data Rows ---
        for row in range(self.table.rowCount()):
            try:
                item_combo = self.table.cellWidget(row, 0)
                if not item_combo or not item_combo.currentText().strip():
                    continue

                row_values = [item_combo.currentText()]
                for col in range(1, self.table.columnCount()):
                    widget = self.table.cellWidget(row, col)
                    if isinstance(widget, QLineEdit):
                        row_values.append(widget.text())
                    elif isinstance(widget, QLabel):
                        row_values.append(widget.text().replace('₹', ''))

                if mode == "Patti":
                    lines.append(header_fmt.format(
                        row_values[0][:5],
                        row_values[1][:3],
                        row_values[2][:3],
                        row_values[3][:5],
                        row_values[4][:3],
                        row_values[5][:7]
                    ))
                elif mode == "Kata":
                    lines.append(header_fmt.format(
                        row_values[0][:5],
                        row_values[1][:3],
                        row_values[2][:3],
                        row_values[3][:5],
                        row_values[4][:3],
                        row_values[5][:7]
                    ))
                elif mode == "Barthe":
                    lines.append(header_fmt.format(
                        row_values[0][:4],
                        row_values[1][:3],
                        row_values[2][:3],
                        row_values[3][:3],
                        row_values[4][:4],
                        row_values[5][:3],
                        row_values[6][:6]
                    ))
            except Exception as e:
                lines.append(f"Format Error: {e}")

        # --- Add Kata Amount if applicable ---
        if mode == "Kata" and hasattr(self, 'kata_amount_entry'):
            kata_amount = validate_float(self.kata_amount_entry.text())
            lines.append(f"Kata Amount: {kata_amount:>8.2f}".rjust(max_width))

        # --- Footer ---
        lines.extend([
            "-" * max_width,
            self.total_label.text().center(max_width),
            "-" * max_width,
            "",
            "ನಾನು ಎಲ್ಲವೂ ಸರಿಯಾಗಿದೆ ಎಂದು ಪರಿಶೀಲಿಸಿದ್ದೇನೆ.".center(max_width),
            "",
            "_" * max_width,
            "Customer Signature".center(max_width),
            "\n\n"  # Extra lines for paper feed
        ])

        return lines

    def save_for_print(self):
        """Prints the generated content to the default printer."""
        try:
            printer_name = win32print.GetDefaultPrinter()
            logging.info(f"Attempting to print to default printer: {printer_name}")
            
            lines = self.generate_print_content()
            print_content = "\n".join(lines)
            
            # First try UTF-8 encoding for Kannada text
            try:
                print_bytes = print_content.encode('utf-8')
            except UnicodeEncodeError:
                # If UTF-8 fails, try UTF-16
                try:
                    print_bytes = print_content.encode('utf-16')
                except UnicodeEncodeError:
                    # If both fail, fall back to cp437 with replacement
                    print_bytes = print_content.encode('cp437', errors='replace')

            # Use win32print for direct RAW printing
            hPrinter = win32print.OpenPrinter(printer_name)
            try:
                hJob = win32print.StartDocPrinter(hPrinter, 1, ("Invoice", None, "RAW"))
                try:
                    win32print.StartPagePrinter(hPrinter)
                    win32print.WritePrinter(hPrinter, print_bytes)
                    win32print.EndPagePrinter(hPrinter)
                finally:
                    win32print.EndDocPrinter(hPrinter)
            finally:
                win32print.ClosePrinter(hPrinter)

            logging.info("Invoice successfully sent to printer.")
            QMessageBox.information(self, "Success", "Invoice sent to printer!")

        except Exception as e:
            error_msg = f"Error printing invoice: {str(e)}"
            logging.error(error_msg)
            logging.error("Printer encoding error - trying to print Kannada text")
            QMessageBox.critical(self, "Print Error",
                f"Could not print to {printer_name}.\n" \
                f"Check if your printer supports Kannada text.\n\nError: {e}")

    def setup_kata_field(self):
        # Clear existing kata field
        for i in reversed(range(self.kata_layout.count())): 
            self.kata_layout.itemAt(i).widget().setParent(None)
        
        if self.current_mode == "Kata":
            kata_label = QLabel("Kata:")
            kata_label.setFont(LABEL_FONT)
            self.kata_amount_entry = QLineEdit()
            self.kata_amount_entry.setFont(ENTRY_FONT)
            self.kata_amount_entry.setFixedWidth(120)
            self.kata_amount_entry.setText("0")
            self.kata_amount_entry.textChanged.connect(self.update_amounts)
            
            self.kata_layout.addWidget(kata_label)
            self.kata_layout.addWidget(self.kata_amount_entry)

    def save_to_excel(self):
            try:
                # Get user's Documents directory
                home_dir = os.path.expanduser("~")
                documents_path = os.path.join(home_dir, "Documents")
                os.makedirs(documents_path, exist_ok=True)
    
                # Create filename based on current date
                date_str = datetime.now().strftime('%Y-%m-%d')
                base_filename = f"Invoice_{date_str}.xlsx"
                full_save_path = os.path.join(documents_path, base_filename)
    
                # Get invoice data
                customer = self.customer_entry.text().strip() or "Unknown Customer"
                mode = self.current_mode
    
                # Get headers
                headers = [self.table.horizontalHeaderItem(i).text() 
                          for i in range(self.table.columnCount())]
    
                # Get row data
                data_rows = []
                for row in range(self.table.rowCount()):
                    row_values = []
                    item_combo = self.table.cellWidget(row, 0)
                    if item_combo and item_combo.currentText().strip():
                        row_values.append(item_combo.currentText())
                        for col in range(1, self.table.columnCount()):
                            widget = self.table.cellWidget(row, col)
                            if isinstance(widget, QLineEdit):
                                row_values.append(widget.text())
                            elif isinstance(widget, QLabel):
                                row_values.append(widget.text().replace('₹', ''))
                        data_rows.append(row_values)
    
                if not data_rows:
                    QMessageBox.warning(self, "No Data", "No data entered to save.")
                    return
    
                # Excel writing logic
                try:
                    if os.path.exists(full_save_path):
                        wb = load_workbook(full_save_path)
                    else:
                        wb = Workbook()
    
                    # Get or create sheet
                    if mode in wb.sheetnames:
                        ws = wb[mode]
                    else:
                        if len(wb.sheetnames) > 0:
                            ws = wb.create_sheet(title=mode)
                        else:
                            ws = wb.active
                            ws.title = mode
    
                    # Clear existing content
                    ws.delete_rows(1, ws.max_row)
    
                    # Write headers and data
                    ws.append(["Timestamp", "Customer"] + headers)
                    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    for row in data_rows:
                        ws.append([timestamp, customer] + row)
    
                    # Save workbook
                    wb.save(full_save_path)
                    QMessageBox.information(self, "Saved", 
                        f"Invoice data saved to:\n{full_save_path}\n(Sheet: {mode})")
    
                except PermissionError:
                    QMessageBox.critical(self, "Permission Error",
                        f"Cannot save '{base_filename}'.\n" \
                        f"The file might be open in Excel.\n\nLocation: {documents_path}")
                except Exception as e:
                    QMessageBox.critical(self, "Save Error",
                        f"Error saving Excel file to:\n{full_save_path}\n\nError: {str(e)}")
    
            except Exception as e:
                QMessageBox.critical(self, "Error",
                    f"Unexpected error during save operation: {str(e)}")
    
   

class PrintPreviewDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.setWindowTitle("Print Preview")
        self.setModal(True)
        self.resize(450, 600)

        # Center the dialog relative to the parent
        if parent:
            geometry = self.geometry()
            center = parent.geometry().center()
            geometry.moveCenter(center)
            self.setGeometry(geometry)

        self.setup_ui()

    def setup_ui(self):
        self.setStyleSheet("""
            QDialog {
                background-color: #f8f9fa;
            }
            QTextEdit {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 10px;
                selection-background-color: #e9ecef;
            }
            QPushButton {
                padding: 10px 20px;
                border-radius: 5px;
                font-weight: bold;
                color: white;
            }
            QPushButton#print_preview { 
                background-color: #0d6efd;
            }
            QPushButton#print_preview:hover {
                background-color: #0b5ed7;
            }
            QPushButton#close {
                background-color: #6c757d;
            }
            QPushButton#close:hover {
                background-color: #5c636a;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Preview text area
        preview_frame = QFrame()
        preview_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 8px;
            }
        """)
        preview_layout = QVBoxLayout(preview_frame)
        preview_layout.setContentsMargins(1, 1, 1, 1)

        self.preview_text = QTextEdit()
        self.preview_text.setFont(QFont("Courier New", 11))
        self.preview_text.setReadOnly(True)
        preview_layout.addWidget(self.preview_text)
        layout.addWidget(preview_frame)

        # Button frame
        button_frame = QFrame()
        button_layout = QHBoxLayout(button_frame)
        button_layout.setSpacing(15)

        print_btn = QPushButton("Print")
        print_btn.setObjectName("print_preview")
        clear_btn = QPushButton("Delete Row")
        clear_btn.setObjectName("delete")
        clear_btn.clicked.connect(self.delete_row)
        button_layout.insertWidget(1, clear_btn)
        close_btn = QPushButton("Close")
        close_btn.setObjectName("close")

        for btn in [print_btn, close_btn]:
            btn.setFont(QFont("Segoe UI", 11, QFont.Bold))
            btn.setMinimumWidth(120)
            btn.setCursor(Qt.PointingHandCursor)
            button_layout.addWidget(btn)

        button_layout.addStretch()

        print_btn.clicked.connect(self.print_invoice)
        close_btn.clicked.connect(self.close)

        layout.addWidget(button_frame)

        # Show preview content
        self.show_preview()

    def show_preview(self):
        try:
            lines = self.parent.generate_print_content()
            preview_content = "\n".join(lines[:-1])  # Exclude printer cut command
            self.preview_text.setPlainText(preview_content)
        except Exception as e:
            QMessageBox.critical(self, "Preview Error",
                f"Error generating preview: {str(e)}")
            self.close()

    def print_invoice(self):
        self.close()
        self.parent.save_for_print()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = InvoiceWindow()
    window.show()
    sys.exit(app.exec())